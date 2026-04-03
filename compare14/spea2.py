#!/usr/bin/env python
# -*- coding: utf-8 -*-

import math
import random
import pathlib
import time
from copy import deepcopy
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional
from pathlib import Path as FSPath
import json as _json

import requests
import json
import numpy as np
import pandas as pd

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401


# ════════════════════════════════════════════════════════
# Global settings
# ════════════════════════════════════════════════════════

TIME_BUCKET_H = 24.0

CHINA_REGIONS   = {"CN"}
EUROPE_REGIONS  = {"EE", "WE"}
TRANSIT_REGIONS = {"KZ", "KG", "UZ", "RU", "BY"}

CORRIDOR_ORDER: Dict[str, int] = {"CN": 0, "CA": 1, "RU": 2, "EE": 3, "WE": 4}

CHINA_BORDER_NODES: set = {"Erenhot", "Manzhouli", "Khorgos", "Lianyungang",
                            "Chongqing", "Yiwu"}

NODE_GROUP: Dict[str, str] = {}

HARD_TIME_WINDOW = False

PEN_MISS_TT            = 5e7
PEN_MISS_ALLOC         = 1e9
PEN_CAP_EXCESS_PER_TEU = 5e7

WAITING_COST_PER_TEU_HOUR_DEFAULT  = 0.8
WAIT_EMISSION_gCO2_per_TEU_H_DEFAULT = 0.0

W_ADD  = 0.25
W_DEL  = 0.25
W_MOD  = 0.25
W_MODE = 0.25
OPS    = ["add", "del", "mod", "mode"]

# Fixed mutation operator weights (uniform — adaptive roulette removed)

CROSSOVER_RATE  = 0.9
MUTATION_RATE   = 0.1

PATHS_TOPK_PER_CRITERION = 15
PATH_LIB_CAP_TOTAL       = 45
DFS_MAX_PATHS_PER_OD     = 150
å
CROSSOVER_SEGMENT_PROB = 0.50

MIN_FEASIBLE_SOLUTIONS       = 10
FEASIBLE_BOOST_ROUNDS        = 20
FEASIBLE_BOOST_MUTATION_RATE = 0.60
FEASIBLE_BOOST_TOPK_PARENTS  = 10

HV_EVERY   = 5
HV_SAMPLES = 2000
METRIC_EVERY = 5

PSTAR_TAIL_GENS  = 30
PSTAR_CAP_PER_GEN = 40
PSTAR_MAX_TOTAL   = 50000

HV_REF_NORM = (1.2, 1.2, 1.2)
HV_MC_SEED  = 12345

DEFAULT_PENALTY_PER_TEU_H = 65.0

# SPEA2-specific
ARCHIVE_SIZE_RATIO = 1.0   # archive_size = pop_size * ratio


# ════════════════════════════════════════════════════════
# [A] Scenario Table  (Table V)
# ════════════════════════════════════════════════════════

SCENARIO_TABLE = {
    "S0":  {"K": 20, "alpha": 1.0},
    "S1":  {"K": 10, "alpha": 0.5},
    "S2":  {"K": 10, "alpha": 1.0},
    "S3":  {"K": 10, "alpha": 2.0},
    "S4":  {"K": 20, "alpha": 0.5},
    "S5":  {"K": 20, "alpha": 2.0},
    "S6":  {"K": 30, "alpha": 0.5},
    "S7":  {"K": 30, "alpha": 1.0},
    "S8":  {"K": 30, "alpha": 2.0},
    "S9":  {"K": 50, "alpha": 0.5},
    "S10": {"K": 50, "alpha": 1.0},
    "S11": {"K": 50, "alpha": 2.0},
}


# ════════════════════════════════════════════════════════
# Corridor constraint helpers
# ════════════════════════════════════════════════════════

def china_border_monotone_ok(nodes: List[str], node_region: Dict[str, str]) -> bool:
    passed_border = False
    left_china    = False
    start_is_border = (len(nodes) > 0 and nodes[0] in CHINA_BORDER_NODES)
    for i, n in enumerate(nodes):
        r        = str(node_region.get(n, "")).strip()
        in_china  = (r in CHINA_REGIONS)
        is_border = (n in CHINA_BORDER_NODES)
        if left_china and in_china:
            return False
        if in_china:
            if passed_border:
                return False
            if is_border and not (i == 0 and start_is_border):
                passed_border = True
        else:
            left_china = True
    return True


def region_monotone_ok(nodes: List[str], node_region: Dict[str, str]) -> bool:
    max_level = -1
    for n in nodes:
        grp   = NODE_GROUP.get(n, "")
        level = CORRIDOR_ORDER.get(grp, -1)
        if level < 0:
            continue
        if level < max_level:
            return False
        max_level = level
    return True


# ════════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════════

def normalize_mode(mode_raw: str) -> str:
    m = str(mode_raw).strip().lower()
    if m in {"railway", "rail"}:  return "rail"
    if m in {"road", "truck"}:    return "road"
    if m in {"water", "ship", "sea"}: return "water"
    return m


def safe_float(x, default=0.0) -> float:
    try:
        if pd.isna(x): return default
    except Exception:
        pass
    try:
        return float(x)
    except Exception:
        return default


def parse_distance_km(x) -> float:
    s       = str(x)
    cleaned = "".join(ch for ch in s if (ch.isdigit() or ch == "."))
    return float(cleaned) if cleaned else 0.0


def norm_region(x: str) -> str:
    s  = str(x).strip()
    if not s or s.lower() in {"nan", "none", ""}: return ""
    sl = s.lower()
    if sl in {"china", "prc", "chn"}:              return "CN"
    if sl in {"we", "west europe", "western europe"}: return "WE"
    if sl in {"ee", "east europe", "eastern europe"}: return "EE"
    return s.upper()


def unique_objective_tuples(objs, tol=1e-9):
    out = []
    for o in objs:
        dup = any(all(abs(o[i] - p[i]) <= tol for i in range(3)) for p in out)
        if not dup:
            out.append(o)
    return out


def _is_bad_text_token(s: str) -> bool:
    if s is None: return True
    t = str(s).strip()
    return t == "" or t.startswith("...")


def _ffill_nan(arr: np.ndarray) -> np.ndarray:
    x = np.array(arr, dtype=float).copy()
    if x.size == 0: return x
    finite_idx = np.where(np.isfinite(x))[0]
    if finite_idx.size == 0: return x
    x[~np.isfinite(x)] = np.nan
    first = finite_idx[0]
    if first > 0:
        x[:first] = x[first]
    for i in range(1, len(x)):
        if np.isnan(x[i]) and np.isfinite(x[i - 1]):
            x[i] = x[i - 1]
    return x


def _finite_points_array(pts):
    if not pts:
        return np.empty((0, 3), dtype=float)
    arr = np.array(pts, dtype=float)
    if arr.ndim != 2 or arr.shape[1] != 3:
        return np.empty((0, 3), dtype=float)
    return arr[np.all(np.isfinite(arr), axis=1)]


# ════════════════════════════════════════════════════════
# Data structures
# ════════════════════════════════════════════════════════

@dataclass
class Arc:
    from_node: str
    to_node: str
    mode: str
    distance: float
    capacity: float
    cost_per_teu_km: float
    emission_per_teu_km: float
    speed_kmh: float
    from_region: str = ""
    to_region:   str = ""


@dataclass
class TimetableEntry:
    from_node: str
    to_node:   str
    mode:      str
    frequency_per_week: float
    first_departure_hour: float
    headway_hours: float


@dataclass
class Batch:
    batch_id:  int
    origin:    str
    destination: str
    quantity:  float
    ET:        float
    LT:        float
    penalty_per_teu_h: float = DEFAULT_PENALTY_PER_TEU_H


@dataclass
class Path:
    path_id:   int
    origin:    str
    destination: str
    nodes:  List[str]
    modes:  List[str]
    arcs:   List[Arc]
    base_cost_per_teu:     float
    base_emission_per_teu: float
    base_travel_time_h:    float

    def __eq__(self, other):
        if not isinstance(other, Path): return NotImplemented
        return self.nodes == other.nodes and self.modes == other.modes

    def __hash__(self):
        return hash((tuple(self.nodes), tuple(self.modes)))


@dataclass
class PathAllocation:
    path:  Path
    share: float

    def __repr__(self):
        chain = ""
        for i, node in enumerate(self.path.nodes[:-1]):
            chain += f"{node}--({self.path.modes[i]})-->"
        chain += self.path.nodes[-1]
        return f"\n    {{ Structure: [{chain}], Share: {self.share:.2%} }}"


@dataclass(eq=False)
class Individual:
    od_allocations: Dict[Tuple[str, str, int], List[PathAllocation]] = field(default_factory=dict)
    objectives:     Tuple[float, float, float] = (float("inf"), float("inf"), float("inf"))
    penalty:        float = 0.0
    feasible:       bool  = False
    feasible_hard:  bool  = False
    vio_breakdown:  Dict[str, float] = field(default_factory=dict)


# ════════════════════════════════════════════════════════
# Merge & normalise shares
# ════════════════════════════════════════════════════════

def merge_and_normalize(allocs: List[PathAllocation]) -> List[PathAllocation]:
    if not allocs: return []
    merged: Dict[Path, float] = {}
    for a in allocs:
        merged[a.path] = merged.get(a.path, 0.0) + float(a.share)
    unique_allocs = [PathAllocation(path=p, share=s) for p, s in merged.items()]
    total = sum(a.share for a in unique_allocs)
    if total <= 1e-12:
        avg = 1.0 / max(1, len(unique_allocs))
        for a in unique_allocs: a.share = avg
    else:
        for a in unique_allocs: a.share /= total
    filtered = [a for a in unique_allocs if a.share > 0.05]
    if not filtered:
        best = max(unique_allocs, key=lambda a: a.share)
        best.share = 1.0
        return [best]
    total2 = sum(a.share for a in filtered)
    if abs(total2 - 1.0) > 1e-9:
        for a in filtered: a.share /= total2
    return filtered


# ════════════════════════════════════════════════════════
# Load data  (identical to NSGA-II version)
# ════════════════════════════════════════════════════════

def load_carbon_tax_map(xls):
    out = {}
    if "Carbon_Tax" not in xls.sheet_names: return out
    try:
        df = pd.read_excel(xls, "Carbon_Tax")
        rc = next((c for c in ["Region","region","RegionCode"] if c in df.columns), None)
        tc = next((c for c in ["CarbonTax_$_per_tCO2","CarbonTax","CT","Tax"] if c in df.columns), None)
        if rc and tc:
            for _, row in df.iterrows():
                r = str(row.get(rc,"")).strip()
                if r: out[r] = safe_float(row.get(tc), default=0.0)
        print(f"[INFO] Loaded carbon tax for {len(out)} regions.")
    except Exception as e:
        print(f"[WARN] Failed to read Carbon_Tax ({e}).")
    return out


def load_border_delay_map(xls):
    out = {}
    if "Node_Border" not in xls.sheet_names: return out
    try:
        nb = pd.read_excel(xls, "Node_Border")
        nc = next((c for c in ["EnglishName","NodeEN"] if c in nb.columns), None)
        mc = next((c for c in ["Mode","mode"] if c in nb.columns), None)
        dc = next((c for c in ["BorderDelay_h","Delay_h","BD"] if c in nb.columns), None)
        if nc and mc and dc:
            for _, row in nb.iterrows():
                n = str(row.get(nc,"")).strip()
                m = normalize_mode(row.get(mc,""))
                if n and m:
                    out[(n, m)] = safe_float(row.get(dc), default=0.0)
        print(f"[INFO] Loaded border delay map: {len(out)} entries.")
    except Exception as e:
        print(f"[WARN] Failed to load border delay ({e}).")
    return out


def load_emission_factor_map(xls):
    out = {}
    if "Emission_Factors" not in xls.sheet_names: return out
    try:
        df  = pd.read_excel(xls, "Emission_Factors")
        mc  = next((c for c in ["Mode","mode"] if c in df.columns), None)
        rc  = next((c for c in ["Region","region","RegionCode"] if c in df.columns), None)
        efc = next((c for c in ["EmissionFactor","Emission_gCO2_per_TEU_km","EF","value"] if c in df.columns), None)
        if mc and rc and efc:
            for _, row in df.iterrows():
                m = normalize_mode(row.get(mc,""))
                r = str(row.get(rc,"")).strip()
                if m and r: out[(m, r)] = safe_float(row.get(efc), default=0.0)
        print(f"[INFO] Loaded emission factors for {len(out)} (mode, region) pairs.")
    except Exception as e:
        print(f"[WARN] Failed to read Emission_Factors ({e}).")
    return out


def load_mode_speeds(xls):
    out = {}
    if "Mode_Speeds" not in xls.sheet_names: return out
    try:
        df  = pd.read_excel(xls, "Mode_Speeds")
        mc  = next((c for c in ["Mode","mode"] if c in df.columns), None)
        spc = next((c for c in ["Speed_kmh","speed_kmh","Speed"] if c in df.columns), None)
        if mc and spc:
            for _, row in df.iterrows():
                m = normalize_mode(row.get(mc,""))
                if m: out[m] = safe_float(row.get(spc), default=0.0)
        print(f"[INFO] Loaded mode speeds: {out}")
    except Exception as e:
        print(f"[WARN] Failed to read Mode_Speeds ({e}).")
    return out


def load_transshipment_map(xls):
    out = {}
    if "Transshipment" not in xls.sheet_names: return out
    try:
        df   = pd.read_excel(xls, "Transshipment")
        ndc  = next((c for c in ["Node","NodeEN","EnglishName"] if c in df.columns), None)
        imc  = next((c for c in ["InMode","FromMode","mode_in"] if c in df.columns), None)
        omc  = next((c for c in ["OutMode","ToMode","mode_out"] if c in df.columns), None)
        cstc = next((c for c in ["TransCost","Cost","trans_cost","Cost_per_TEU"] if c in df.columns), None)
        tmc  = next((c for c in ["TransTime_h","Time_h","trans_time_h","Time"] if c in df.columns), None)
        if ndc and imc and omc:
            for _, row in df.iterrows():
                node    = str(row.get(ndc,"")).strip()
                in_mode = normalize_mode(row.get(imc,""))
                out_mode= normalize_mode(row.get(omc,""))
                if node and in_mode and out_mode:
                    out[(node, in_mode, out_mode)] = {
                        "cost_per_teu": safe_float(row.get(cstc), default=0.0) if cstc else 0.0,
                        "time_h":       safe_float(row.get(tmc),  default=0.0) if tmc  else 0.0,
                    }
        print(f"[INFO] Loaded transshipment entries: {len(out)}")
    except Exception as e:
        print(f"[WARN] Failed to read Transshipment ({e}).")
    return out


def load_waiting_params(xls):
    wc = WAITING_COST_PER_TEU_HOUR_DEFAULT
    we = WAIT_EMISSION_gCO2_per_TEU_H_DEFAULT
    if "Waiting_Costs" not in xls.sheet_names: return wc, we
    try:
        df = pd.read_excel(xls, "Waiting_Costs")
        def pick(colnames, default):
            for c in colnames:
                if c in df.columns:
                    vals = df[c].dropna().tolist()
                    if vals: return safe_float(vals[0], default=default)
            return default
        wc = pick(["WaitingCost_per_TEU_h","WaitCost_per_TEU_h"], wc)
        we = pick(["WaitEmission_gCO2_per_TEU_h","WaitingEmission_gCO2_per_TEU_h"], we)
        print(f"[INFO] Loaded waiting params: cost={wc}, emission={we}")
    except Exception as e:
        print(f"[WARN] Failed to read Waiting_Costs ({e}).")
    return wc, we


def load_network_from_extended(filename: str):
    global CHINA_BORDER_NODES, NODE_GROUP
    xls = pd.ExcelFile(filename)

    carbon_tax_map     = load_carbon_tax_map(xls)
    emission_factor_map= load_emission_factor_map(xls)
    mode_speeds_map    = load_mode_speeds(xls)
    trans_map          = load_transshipment_map(xls)
    border_delay_map   = load_border_delay_map(xls)

    nodes_df   = pd.read_excel(xls, "Nodes")
    node_names = nodes_df["EnglishName"].astype(str).str.strip().tolist()

    node_region = {
        str(name).strip(): norm_region(reg)
        for name, reg in zip(nodes_df["EnglishName"], nodes_df["Region"])
    }

    if "RegionGroup" in nodes_df.columns:
        NODE_GROUP = {
            str(name).strip(): str(grp).strip()
            for name, grp in zip(nodes_df["EnglishName"], nodes_df["RegionGroup"])
            if str(grp).strip() not in ("", "nan", "None")
        }

    if "Node_Border" in xls.sheet_names:
        try:
            nb_df = pd.read_excel(xls, "Node_Border")
            loaded_borders = set()
            for _, row in nb_df.iterrows():
                region_val = str(row.get("Region","")).strip()
                is_border  = safe_float(row.get("IsBorderNode", 0), default=0.0) == 1.0
                if region_val == "CN" and is_border:
                    node_name = str(row.get("EnglishName","")).strip()
                    if node_name: loaded_borders.add(node_name)
            if loaded_borders:
                CHINA_BORDER_NODES = loaded_borders
                print(f"[INFO] Loaded China border nodes ({len(CHINA_BORDER_NODES)}): {sorted(CHINA_BORDER_NODES)}")
        except Exception as e:
            print(f"[WARN] Failed to load Node_Border: {e}. Using defaults.")
    else:
        print(f"[INFO] Node_Border sheet not found. Using defaults: {sorted(CHINA_BORDER_NODES)}")

    CHINA_BORDER_NODES.update({"Ningbo","Shanghai"})

    node_hold_cost: Dict[str, float] = {}
    node_proc_cost: Dict[str, float] = {}
    for _, row in nodes_df.iterrows():
        n = str(row.get("EnglishName","")).strip()
        node_hold_cost[n] = safe_float(row.get("HoldCost_per_TEU_h"), default=WAITING_COST_PER_TEU_HOUR_DEFAULT)
        node_proc_cost[n] = safe_float(row.get("ProcCost_per_TEU_h"), default=0.0)

    SEAPORT_NODES = {"Ningbo","Shanghai"}
    for n in SEAPORT_NODES:
        if n not in node_region:
            node_region[n] = "CN"
            NODE_GROUP[n]  = "CN"
            if n not in node_names: node_names.append(n)

    waiting_cost_per_teu_h, wait_emis_g_per_teu_h = load_waiting_params(xls)

    arcs_df   = pd.read_excel(xls, "Arcs_All")
    arcs: List[Arc] = []
    cost_cols = ["Cost_$_per_km","Cost_per_km","Cost"]
    emis_cols = ["Emission_gCO2_per_tkm","Emission_gCO2_per_TEU_km","EmissionFactor","Emission"]

    for _, row in arcs_df.iterrows():
        mode  = normalize_mode(row.get("Mode","road"))
        speed = {"road": 75.0, "water": 30.0}.get(mode, 50.0)
        if mode in mode_speeds_map and mode_speeds_map[mode] > 0:
            speed = mode_speeds_map[mode]

        origin = str(row.get("OriginEN","")).strip()
        dest   = str(row.get("DestEN","")).strip()
        if _is_bad_text_token(origin) or _is_bad_text_token(dest): continue

        from_region = str(node_region.get(origin,"")).strip()
        to_region   = str(node_region.get(dest,"")).strip()
        distance    = parse_distance_km(row.get("Distance_km", 0.0))

        if "Capacity_TEUday" in arcs_df.columns and not pd.isna(row.get("Capacity_TEUday", np.nan)):
            capacity = safe_float(row.get("Capacity_TEUday"), default=1e9)
        elif "Capacity_TEUh" in arcs_df.columns and not pd.isna(row.get("Capacity_TEUh", np.nan)):
            capacity = safe_float(row.get("Capacity_TEUh"), default=1e9) * 24.0
        else:
            capacity = 1e9

        cpkm = 0.0
        for c in cost_cols:
            if c in arcs_df.columns:
                val = safe_float(row.get(c), default=None)
                if val is not None and val > 0: cpkm = val; break
        if cpkm <= 1e-9: cpkm = 0.5

        epkm = 0.0
        for c in emis_cols:
            if c in arcs_df.columns:
                epkm = safe_float(row.get(c), default=0.0); break
        if (mode, from_region) in emission_factor_map:
            epkm = emission_factor_map[(mode, from_region)]

        arcs.append(Arc(
            from_node=origin, to_node=dest, mode=mode,
            distance=distance, capacity=capacity,
            cost_per_teu_km=cpkm, emission_per_teu_km=epkm,
            speed_kmh=speed, from_region=from_region, to_region=to_region
        ))

    tdf = pd.read_excel(xls, "Timetable")
    timetables: List[TimetableEntry] = []
    for _, row in tdf.iterrows():
        origin    = str(row.get("OriginEN","")).strip()
        dest      = str(row.get("DestEN","")).strip()
        mode_norm = normalize_mode(row.get("Mode",""))
        if _is_bad_text_token(origin) or _is_bad_text_token(dest): continue
        if mode_norm not in {"road","rail","water"}: continue
        freq   = safe_float(row.get("Frequency_per_week"), default=1.0)
        hd_raw = row.get("Headway_Hours", np.nan)
        hd     = 168.0 / max(freq, 1.0) if pd.isna(hd_raw) else safe_float(hd_raw, default=168.0)
        v      = row.get("FirstDepartureHour", np.nan)
        fd     = 0.0
        if not pd.isna(v):
            try:
                s  = str(v).strip()
                fd = float(s.split(":")[0]) if ":" in s else float(s)
            except Exception:
                fd = 0.0
        timetables.append(TimetableEntry(
            from_node=origin, to_node=dest, mode=mode_norm,
            frequency_per_week=freq, first_departure_hour=fd, headway_hours=hd
        ))

    bdf     = pd.read_excel(xls, "Batches")
    bdf     = augment_batches_to_20(bdf, node_region=node_region, random_seed=2026)
    batches: List[Batch] = []
    for _, row in bdf.iterrows():
        origin = str(row.get("OriginEN","")).strip()
        dest   = str(row.get("DestEN","")).strip()
        if node_region.get(origin) in CHINA_REGIONS and node_region.get(dest) in EUROPE_REGIONS:
            batches.append(Batch(
                batch_id=int(row.get("BatchID", 0)),
                origin=origin, destination=dest,
                quantity=safe_float(row.get("QuantityTEU"), default=0.0),
                ET=safe_float(row.get("ET"), default=0.0),
                LT=safe_float(row.get("LT"), default=0.0),
                penalty_per_teu_h=safe_float(row.get("PenaltyCost_per_TEU_h"),
                                             default=DEFAULT_PENALTY_PER_TEU_H)
            ))

    print(f"[INFO] Batches loaded: {len(batches)}")
    print(f"[INFO] Node capacity constraints: DISABLED")
    return (
        node_names, node_region,
        node_hold_cost, node_proc_cost,
        arcs, timetables, batches,
        waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
        carbon_tax_map, emission_factor_map, mode_speeds_map,
        trans_map, border_delay_map
    )


def build_graph(arcs):
    g = {}
    for a in arcs:
        g.setdefault(a.from_node, []).append((a.to_node, a))
    return g


def build_timetable_dict(timetables):
    tt = {}
    for t in timetables:
        tt.setdefault((t.from_node, t.to_node, t.mode), []).append(t)
    return tt


def build_arc_lookup(arcs):
    mp = {}
    for a in arcs:
        k = (a.from_node, a.to_node, a.mode)
        if k not in mp: mp[k] = a
    return mp


# ════════════════════════════════════════════════════════
# Path library
# ════════════════════════════════════════════════════════

def random_dfs_paths(graph, origin, dest, node_region,
                     max_len=12, max_paths=200, timeout_sec=8.0):
    deadline = time.time() + timeout_sec
    paths, found_set = [], set()
    attempts, max_attempts = 0, max_paths * 20
    while len(paths) < max_paths and attempts < max_attempts:
        if time.time() > deadline: break
        attempts += 1
        node, cur_arcs, visited, cur_nodes, ok = origin, [], {origin}, [origin], True
        for _ in range(max_len):
            if node == dest: break
            neighbors = list(graph.get(node, []))
            if not neighbors: ok = False; break
            random.shuffle(neighbors)
            moved = False
            for nxt, arc in neighbors:
                if nxt in visited: continue
                new_nodes = cur_nodes + [nxt]
                if not china_border_monotone_ok(new_nodes, node_region): continue
                if not region_monotone_ok(new_nodes, node_region): continue
                cur_arcs.append(arc); visited.add(nxt)
                cur_nodes.append(nxt); node = nxt; moved = True; break
            if not moved: ok = False; break
        if ok and node == dest and cur_arcs:
            key = tuple(cur_nodes)
            if key not in found_set:
                found_set.add(key); paths.append(cur_arcs)
    return paths


def repair_arc_seq_with_road_fallback(arc_seq, tt_dict, arc_lookup):
    new_seq = []
    for arc in arc_seq:
        if arc.mode == "road": new_seq.append(arc); continue
        if tt_dict.get((arc.from_node, arc.to_node, arc.mode), []):
            new_seq.append(arc); continue
        k_road = (arc.from_node, arc.to_node, "road")
        if k_road in arc_lookup: new_seq.append(arc_lookup[k_road])
        else: return None
    return new_seq


def select_topk_by_cost_time_emis(paths, k=30, cap_total=90):
    if not paths: return []
    by_cost = sorted(paths, key=lambda p: p.base_cost_per_teu)
    by_time = sorted(paths, key=lambda p: p.base_travel_time_h)
    by_emis = sorted(paths, key=lambda p: p.base_emission_per_teu)
    picked, used = [], set()
    for lst in [by_cost, by_time, by_emis]:
        for p in lst[:k]:
            if p not in used: picked.append(p); used.add(p)
    return picked[:cap_total] if cap_total else picked


def build_path_library(node_names, node_region, arcs, batches, tt_dict, arc_lookup):
    graph    = build_graph(arcs)
    path_lib = {}
    next_pid = 0
    for b in batches:
        od = (b.origin, b.destination)
        if od in path_lib: continue
        arc_paths = random_dfs_paths(graph, b.origin, b.destination,
                                     node_region=node_region, max_len=12,
                                     max_paths=DFS_MAX_PATHS_PER_OD)
        paths_od = []
        for arc_seq in arc_paths:
            repaired = repair_arc_seq_with_road_fallback(arc_seq, tt_dict, arc_lookup)
            if repaired is None: continue
            nodes = [repaired[0].from_node] + [a.to_node for a in repaired]
            if len(set(nodes)) != len(nodes): continue
            if not region_monotone_ok(nodes, node_region): continue
            if not china_border_monotone_ok(nodes, node_region): continue
            modes = [a.mode for a in repaired]
            paths_od.append(Path(
                path_id=next_pid, origin=b.origin, destination=b.destination,
                nodes=nodes, modes=modes, arcs=repaired,
                base_cost_per_teu=sum(a.cost_per_teu_km * a.distance for a in repaired),
                base_emission_per_teu=sum(a.emission_per_teu_km * a.distance for a in repaired),
                base_travel_time_h=sum(a.distance / max(a.speed_kmh, 1.0) for a in repaired),
            ))
            next_pid += 1
        if paths_od:
            path_lib[od] = select_topk_by_cost_time_emis(
                paths_od, k=PATHS_TOPK_PER_CRITERION, cap_total=PATH_LIB_CAP_TOTAL)

    removed = 0
    for od in list(path_lib.keys()):
        before = len(path_lib[od])
        path_lib[od] = [p for p in path_lib[od]
                        if china_border_monotone_ok(p.nodes, node_region)]
        removed += before - len(path_lib[od])
        if not path_lib[od]: del path_lib[od]
    if removed: print(f"[WARN] Post-filter removed {removed} paths.")
    else:       print("[INFO] All paths pass border monotonicity. ✅")
    return path_lib


def sanity_check_path_lib(batches, path_lib):
    missing = [(b.batch_id, (b.origin, b.destination))
               for b in batches if not path_lib.get((b.origin, b.destination), [])]
    if missing:
        for bid, od in missing[:20]:
            print(f"[SANITY] ❌ missing paths Batch {bid} OD={od}")
        raise RuntimeError("Path library missing some ODs.")
    print("[SANITY] ✅ All batches have paths.")


def repair_missing_allocations(ind, batches, path_lib):
    for b in batches:
        key   = (b.origin, b.destination, b.batch_id)
        if ind.od_allocations.get(key, []): continue
        paths = path_lib.get((b.origin, b.destination), [])
        if paths:
            ind.od_allocations[key] = [PathAllocation(path=paths[0], share=1.0)]


# ════════════════════════════════════════════════════════
# Simulation & evaluation
# ════════════════════════════════════════════════════════

def next_departure_time_programB(t: float, entries: List[TimetableEntry]) -> float:
    best = float("inf")
    for e in entries:
        if t <= e.first_departure_hour:
            dep = e.first_departure_hour
        else:
            waited = t - e.first_departure_hour
            n      = math.ceil(waited / max(e.headway_hours, 1e-6))
            dep    = e.first_departure_hour + n * e.headway_hours
        if dep < best: best = dep
    return best if best < float("inf") else t


def simulate_path_time_capacity(
    path: Path, batch: Batch, flow_teu: float,
    tt_dict: Dict, arc_flow_map: Dict,
    trans_map: Optional[Dict] = None,
    border_delay_map: Optional[Dict] = None,
) -> Tuple[float, List[Tuple[str, float, float]], int]:
    t               = float(batch.ET)
    miss_tt         = 0
    trans_map       = trans_map or {}
    border_delay_map= border_delay_map or {}
    prev_arc        = None
    node_wait_list: List[Tuple[str, float, float]] = []

    for arc in path.arcs:
        cur_node       = arc.from_node
        arc_trans_wait = 0.0

        if prev_arc is not None and prev_arc.mode != arc.mode:
            rec = trans_map.get((cur_node, prev_arc.mode, arc.mode))
            if rec:
                th = safe_float(rec.get("time_h"), default=0.0)
                if th > 0: t += th; arc_trans_wait += th

        if cur_node in CHINA_BORDER_NODES:
            bd = border_delay_map.get((cur_node, arc.mode), 0.0)
            if bd > 0: t += bd; arc_trans_wait += bd

        travel_arc = arc.distance / max(arc.speed_kmh, 1.0)
        entries    = [] if arc.mode == "road" else \
                     tt_dict.get((cur_node, arc.to_node, arc.mode), [])
        if arc.mode != "road" and not entries:
            miss_tt += 1; return float("inf"), [], miss_tt

        dep            = t if not entries else next_departure_time_programB(t, entries)
        arc_sched_wait = max(0.0, dep - t)
        node_wait_list.append((cur_node, arc_sched_wait, arc_trans_wait))

        arr  = dep + travel_arc
        slot = int(dep // 24)
        akey = (cur_node, arc.to_node, arc.mode)
        arc_flow_map[(akey, slot)] = arc_flow_map.get((akey, slot), 0.0) + flow_teu

        t        = arr
        prev_arc = arc

    return (t - batch.ET), node_wait_list, miss_tt


def evaluate_individual(
    ind, batches, arcs, tt_dict,
    waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
    node_hold_cost=None, node_proc_cost=None,
    carbon_tax_map=None, trans_map=None, border_delay_map=None,
):
    node_hold_cost   = node_hold_cost   or {}
    node_proc_cost   = node_proc_cost   or {}
    carbon_tax_map   = carbon_tax_map   or {}
    trans_map        = trans_map        or {}
    border_delay_map = border_delay_map or {}

    total_cost = total_emission_g = makespan = 0.0
    arc_flow_map: Dict = {}
    arc_caps = {(a.from_node, a.to_node, a.mode): a.capacity for a in arcs}

    miss_alloc = miss_tt = 0
    cap_excess = late_teu_h_total = wait_teu_h_total = 0.0
    trans_teu_h_total = trans_cost_total = carbon_cost_total = 0.0

    for b in batches:
        key   = (b.origin, b.destination, b.batch_id)
        allocs= ind.od_allocations.get(key, [])
        if not allocs: miss_alloc += 1; continue

        batch_finish = b.ET
        for alloc in allocs:
            if alloc.share <= 1e-12: continue
            flow = alloc.share * b.quantity
            p    = alloc.path

            travel_time, node_wait_list, mtt = simulate_path_time_capacity(
                p, b, flow, tt_dict, arc_flow_map,
                trans_map=trans_map, border_delay_map=border_delay_map)
            if math.isinf(travel_time): miss_tt += mtt; continue

            total_cost       += p.base_cost_per_teu * flow
            total_emission_g += p.base_emission_per_teu * flow

            cc = sum(
                (arc.emission_per_teu_km * arc.distance * flow / 1e6)
                * float(carbon_tax_map.get(getattr(arc, "from_region",""), 0.0))
                for arc in p.arcs)
            total_cost += cc; carbon_cost_total += cc

            tc = 0.0
            for i in range(len(p.arcs) - 1):
                if p.arcs[i].mode != p.arcs[i+1].mode:
                    node = p.arcs[i+1].from_node
                    rec  = trans_map.get((node, p.arcs[i].mode, p.arcs[i+1].mode), {})
                    tc  += safe_float(rec.get("cost_per_teu"), default=0.0) * flow
            total_cost += tc; trans_cost_total += tc

            for (wnode, sched_h, trans_h) in node_wait_list:
                hold_rate = node_hold_cost.get(wnode, WAITING_COST_PER_TEU_HOUR_DEFAULT)
                proc_rate = node_proc_cost.get(wnode, 0.0)
                if sched_h > 0.0:
                    total_cost       += hold_rate * flow * sched_h
                    total_emission_g += wait_emis_g_per_teu_h * flow * sched_h
                    wait_teu_h_total += flow * sched_h
                if trans_h > 0.0:
                    total_cost         += proc_rate * flow * trans_h
                    trans_teu_h_total  += flow * trans_h

            arrival_time = b.ET + travel_time
            batch_finish = max(batch_finish, arrival_time)
            if arrival_time > b.LT:
                late_h            = flow * (arrival_time - b.LT)
                late_teu_h_total += late_h
                total_cost       += b.penalty_per_teu_h * late_h

        makespan = max(makespan, batch_finish)

    for (akey, slot), sf in arc_flow_map.items():
        cap = arc_caps.get(akey, 1e9)
        if sf > cap: cap_excess += (sf - cap)

    penalty = (PEN_MISS_ALLOC * float(miss_alloc) +
               PEN_MISS_TT    * float(miss_tt)     +
               PEN_CAP_EXCESS_PER_TEU * float(cap_excess))

    ind.objectives    = (float(total_cost), float(total_emission_g), float(makespan))
    ind.penalty       = float(penalty)
    hard_ok           = (miss_alloc == 0 and miss_tt == 0 and cap_excess <= 1e-9)
    ind.feasible_hard = bool(hard_ok)
    ind.feasible      = bool(hard_ok)
    ind.vio_breakdown = {
        "miss_alloc":    float(miss_alloc),
        "miss_tt":       float(miss_tt),
        "cap_excess":    float(cap_excess),
        "late_teu_h":    float(late_teu_h_total),
        "wait_teu_h":    float(wait_teu_h_total),
        "trans_teu_h":   float(trans_teu_h_total),
        "trans_cost":    float(trans_cost_total),
        "carbon_cost":   float(carbon_cost_total),
    }


# ════════════════════════════════════════════════════════
# GA operators  (identical to NSGA-II)
# ════════════════════════════════════════════════════════

def clone_gene(alloc):
    return PathAllocation(path=alloc.path, share=float(alloc.share))


def crossover_structural(ind1, ind2, batches):
    child1, child2 = Individual(), Individual()
    for b in batches:
        key = (b.origin, b.destination, b.batch_id)
        g1  = ind1.od_allocations.get(key, [])
        g2  = ind2.od_allocations.get(key, [])
        if not g1 and not g2: continue
        if not g1:
            child1.od_allocations[key] = [clone_gene(x) for x in g2]
            child2.od_allocations[key] = [clone_gene(x) for x in g2]; continue
        if not g2:
            child1.od_allocations[key] = [clone_gene(x) for x in g1]
            child2.od_allocations[key] = [clone_gene(x) for x in g1]; continue
        cut1, cut2 = random.randint(0, len(g1)), random.randint(0, len(g2))
        c1 = [clone_gene(x) for x in g1[:cut1]] + [clone_gene(x) for x in g2[cut2:]]
        c2 = [clone_gene(x) for x in g2[:cut2]] + [clone_gene(x) for x in g1[cut1:]]
        child1.od_allocations[key] = merge_and_normalize(c1)
        child2.od_allocations[key] = merge_and_normalize(c2)
    return child1, child2


def path_from_arcs(new_arcs, origin, destination, path_id=-1, node_region=None):
    if not new_arcs: return None
    nodes = [new_arcs[0].from_node] + [a.to_node for a in new_arcs]
    if nodes[0] != origin or nodes[-1] != destination: return None
    if len(set(nodes)) != len(nodes): return None
    if node_region is not None:
        if not china_border_monotone_ok(nodes, node_region): return None
        if not region_monotone_ok(nodes, node_region):       return None
    return Path(
        path_id=path_id, origin=origin, destination=destination,
        nodes=nodes, modes=[a.mode for a in new_arcs], arcs=new_arcs,
        base_cost_per_teu=sum(a.cost_per_teu_km * a.distance for a in new_arcs),
        base_emission_per_teu=sum(a.emission_per_teu_km * a.distance for a in new_arcs),
        base_travel_time_h=sum(a.distance / max(a.speed_kmh, 1.0) for a in new_arcs),
    )


def rebuild_path_from_nodes_modes(origin, destination, nodes, modes,
                                   tt_dict, arc_lookup, allow_road_fallback=True):
    if not nodes or len(nodes) < 2 or nodes[0] != origin or nodes[-1] != destination:
        return None
    if len(modes) != len(nodes) - 1 or len(set(nodes)) != len(nodes): return None
    new_arcs = []
    for i in range(len(modes)):
        u, v, m = nodes[i], nodes[i+1], modes[i]
        k = (u, v, m)
        if k not in arc_lookup: return None
        arc = arc_lookup[k]
        if arc.mode != "road" and not tt_dict.get((u, v, arc.mode), []):
            if allow_road_fallback and (u, v, "road") in arc_lookup:
                arc = arc_lookup[(u, v, "road")]
            else: return None
        new_arcs.append(arc)
    return path_from_arcs(new_arcs, origin, destination)


def find_common_internal_nodes(p1, p2):
    return list(set(p1.nodes[1:-1]) & set(p2.nodes[1:-1]))


def perform_single_point_crossover_paths(pA, pB, join_node, tt_dict, arc_lookup):
    if join_node not in pA.nodes or join_node not in pB.nodes: return None
    ia, ib = pA.nodes.index(join_node), pB.nodes.index(join_node)
    return rebuild_path_from_nodes_modes(
        pA.origin, pA.destination,
        pA.nodes[:ia+1] + pB.nodes[ib+1:],
        pA.modes[:ia]   + pB.modes[ib:],
        tt_dict, arc_lookup)


def crossover_common_node(ind1, ind2, batches, tt_dict, arc_lookup):
    child1, child2 = Individual(), Individual()
    for b in batches:
        key = (b.origin, b.destination, b.batch_id)
        g1  = ind1.od_allocations.get(key, [])
        g2  = ind2.od_allocations.get(key, [])
        if not g1 and not g2: continue
        if not g1:
            child1.od_allocations[key] = [clone_gene(x) for x in g2]
            child2.od_allocations[key] = [clone_gene(x) for x in g2]; continue
        if not g2:
            child1.od_allocations[key] = [clone_gene(x) for x in g1]
            child2.od_allocations[key] = [clone_gene(x) for x in g1]; continue
        c1_allocs = [clone_gene(x) for x in g1]
        c2_allocs = [clone_gene(x) for x in g2]
        p1, p2    = random.choice(g1).path, random.choice(g2).path
        common    = find_common_internal_nodes(p1, p2)
        if common:
            join = random.choice(common)
            np1  = perform_single_point_crossover_paths(p1, p2, join, tt_dict, arc_lookup)
            np2  = perform_single_point_crossover_paths(p2, p1, join, tt_dict, arc_lookup)
            if np1: c1_allocs.append(PathAllocation(path=np1, share=0.20))
            if np2: c2_allocs.append(PathAllocation(path=np2, share=0.20))
        child1.od_allocations[key] = merge_and_normalize(c1_allocs)
        child2.od_allocations[key] = merge_and_normalize(c2_allocs)
    return child1, child2


def crossover_hybrid(p1, p2, batches, tt_dict, arc_lookup):
    if random.random() < CROSSOVER_SEGMENT_PROB:
        c1, c2 = crossover_common_node(p1, p2, batches, tt_dict, arc_lookup)
        if c1.od_allocations or c2.od_allocations: return c1, c2
    return crossover_structural(p1, p2, batches)


def random_initial_individual(batches, path_lib, max_paths=3):
    ind = Individual()
    for b in batches:
        paths = path_lib.get((b.origin, b.destination), [])
        if not paths: continue
        k      = random.randint(1, min(max_paths, len(paths)))
        chosen = random.sample(paths, k)
        raw    = [PathAllocation(path=p, share=random.random()) for p in chosen]
        ind.od_allocations[(b.origin, b.destination, b.batch_id)] = merge_and_normalize(raw)
    return ind


def greedy_initial_individual(batches, path_lib):
    ind = Individual()
    for b in batches:
        paths = path_lib.get((b.origin, b.destination), [])
        if not paths: continue
        best  = min(paths, key=lambda p: p.base_travel_time_h)
        ind.od_allocations[(b.origin, b.destination, b.batch_id)] = \
            [PathAllocation(path=best, share=1.0)]
    return ind


# ════════════════════════════════════════════════════════
# LLM辅助种群生成模块
# ════════════════════════════════════════════════════════

DEEPSEEK_API_KEY = "你的新密钥"  # ← 换成你的新密钥
LLM_THETA        = 0.2
LLM_TOP_K_PATHS  = 10


def select_paths_for_llm(paths: List[Path], top_k: int = 10) -> List[Path]:
    if len(paths) <= top_k:
        return paths
    selected = set()
    for p in sorted(paths, key=lambda p: p.base_travel_time_h)[:3]:
        selected.add(p)
    for p in sorted(paths, key=lambda p: p.base_cost_per_teu)[:3]:
        selected.add(p)
    for p in sorted(paths, key=lambda p: p.base_emission_per_teu)[:3]:
        selected.add(p)
    remaining = [p for p in paths if p not in selected]
    random.shuffle(remaining)
    for p in remaining:
        if len(selected) >= top_k:
            break
        selected.add(p)
    return list(selected)


def format_elite_for_batch(archive: List[Individual],
                            orig: str, dest: str, bid: int,
                            n_samples: int = 3) -> str:
    if not archive:
        return "  暂无历史参考方案"
    samples = random.sample(archive, min(n_samples, len(archive)))
    text = ""
    for i, ind in enumerate(samples):
        key = (orig, dest, bid)
        if key not in ind.od_allocations:
            continue
        text += (f"\n  精英解{i+1} "
                 f"(成本={ind.objectives[0]:.0f} "
                 f"碳排放={ind.objectives[1]:.0f} "
                 f"时间={ind.objectives[2]:.1f}h):\n")
        for alloc in ind.od_allocations[key]:
            nodes_str = " → ".join(alloc.path.nodes)
            modes_str = ", ".join(alloc.path.modes)
            text += (f"    路径: {nodes_str} "
                     f"| 方式: {modes_str} "
                     f"| 比例: {alloc.share:.1%}\n")
    return text if text else "  暂无历史参考方案"


def llm_generate_for_one_batch(
    archive:   List[Individual],
    path_lib:  Dict[Tuple[str, str], List[Path]],
    batches:   List[Batch],
    theta:     float = LLM_THETA,
    top_k:     int   = LLM_TOP_K_PATHS
) -> Optional[Tuple[Tuple[str, str, int], List[PathAllocation]]]:

    if random.random() > theta:
        return None
    if len(archive) < 1:
        return None

    batch = random.choice(batches)
    orig  = batch.origin
    dest  = batch.destination
    bid   = batch.batch_id

    all_paths = path_lib.get((orig, dest), [])
    if not all_paths:
        return None

    selected    = select_paths_for_llm(all_paths, top_k=top_k)
    path_lookup = {p.path_id: p for p in selected}

    paths_text = ""
    for p in selected:
        nodes_str = " → ".join(p.nodes)
        modes_str = ", ".join(p.modes)
        paths_text += (
            f"  [ID={p.path_id}] {nodes_str} "
            f"| 方式: {modes_str} "
            f"| 时间: {p.base_travel_time_h:.1f}h "
            f"| 成本: {p.base_cost_per_teu:.0f} "
            f"| 碳排放: {p.base_emission_per_teu:.1f}\n"
        )

    elite_text = format_elite_for_batch(archive, orig, dest, bid)

    prompt = f"""你是多模式货运路径优化专家。
目标：同时优化总运输时间、总成本、总碳排放（三者越小越好）。

当前任务：
  起点: {orig}
  终点: {dest}
  批次ID: {bid}
  货物数量: {batch.quantity:.0f} TEU

可用路径（共{len(selected)}条）：
{paths_text}
历史优质方案中该OD对的分配情况（供参考）：
{elite_text}

请选择1到3条路径并分配比例（share之和必须精确等于1.0）。

严格按以下JSON格式输出，不要任何其他文字：
{{
  "paths": [
    {{"path_id": 路径ID, "share": 比例}}
  ]
}}"""

    try:
        response = requests.post(
            url="https://api.deepseek.com/v1/chat/completions",
            headers={
                "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "model": "deepseek-chat",
                "messages": [{"role": "user", "content": prompt}],
                "temperature": 0.7
            },
            timeout=30
        )

        result_text = response.json()["choices"][0]["message"]["content"]
        result_json = json.loads(result_text)

        total_share = sum(p["share"] for p in result_json["paths"])
        if abs(total_share - 1.0) > 0.05:
            return None
        for p in result_json["paths"]:
            p["share"] /= total_share

        alloc_list = []
        for p_info in result_json["paths"]:
            pid   = p_info["path_id"]
            share = float(p_info["share"])
            if pid in path_lookup and share > 0:
                alloc_list.append(
                    PathAllocation(path=path_lookup[pid], share=share)
                )

        if not alloc_list:
            return None

        print(f"[LLM] 成功: {orig}→{dest} batch{bid}, {len(alloc_list)}条路径")
        return (orig, dest, bid), alloc_list

    except Exception as e:
        print(f"[LLM] 调用失败: {e}")
        return None

def mutate_add(ind, batch, path_lib):
    key  = (batch.origin, batch.destination, batch.batch_id)
    allocs= ind.od_allocations.get(key, [])
    pool  = path_lib.get((batch.origin, batch.destination), [])
    if not pool: return False
    cur   = {a.path for a in allocs}
    cands = [p for p in pool if p not in cur]
    if not cands:
        if allocs:
            allocs[random.randrange(len(allocs))] = PathAllocation(
                path=random.choice(pool), share=0.2)
            ind.od_allocations[key] = merge_and_normalize(allocs)
            return True
        return False
    allocs.append(PathAllocation(path=random.choice(cands), share=0.2))
    ind.od_allocations[key] = merge_and_normalize(allocs)
    return True


def mutate_del(ind, batch):
    key   = (batch.origin, batch.destination, batch.batch_id)
    allocs= ind.od_allocations.get(key, [])
    if len(allocs) <= 1: return False
    allocs.pop(random.randrange(len(allocs)))
    ind.od_allocations[key] = merge_and_normalize(allocs)
    return True


def mutate_mod(ind, batch):
    key   = (batch.origin, batch.destination, batch.batch_id)
    allocs= ind.od_allocations.get(key, [])
    if not allocs: return False
    random.choice(allocs).share *= random.uniform(0.5, 1.5)
    ind.od_allocations[key] = merge_and_normalize(allocs)
    return True


def mutate_mode(ind, batch, tt_dict, arc_lookup, max_trials=20):
    key   = (batch.origin, batch.destination, batch.batch_id)
    allocs= ind.od_allocations.get(key, [])
    if not allocs: return False
    idx      = random.randrange(len(allocs))
    old_alloc= allocs[idx]
    p        = old_alloc.path
    if not p.arcs: return False
    arc_i    = random.randrange(len(p.arcs))
    old_arc  = p.arcs[arc_i]
    u, v     = old_arc.from_node, old_arc.to_node
    for _ in range(max_trials):
        new_mode = random.choice([m for m in ["road","rail","water"] if m != old_arc.mode])
        k_arc    = (u, v, new_mode)
        if k_arc not in arc_lookup: continue
        if new_mode != "road" and not tt_dict.get((u, v, new_mode), []): continue
        new_arcs    = list(p.arcs)
        new_arcs[arc_i] = arc_lookup[k_arc]
        new_path    = path_from_arcs(new_arcs, p.origin, p.destination)
        if new_path is None: continue
        allocs_new  = deepcopy(allocs)
        allocs_new[idx] = PathAllocation(path=new_path, share=old_alloc.share)
        ind.od_allocations[key] = merge_and_normalize(allocs_new)
        return True
    return False


def augment_batches_to_20(bdf, node_region, random_seed=2026):
    df = bdf.copy()
    required_cols = ["BatchID","OriginEN","DestEN","QuantityTEU","ET","LT"]
    if any(c not in df.columns for c in required_cols) or len(df) >= 20:
        return df
    china_nodes  = [n for n, r in node_region.items()
                    if r in CHINA_REGIONS and n not in CHINA_BORDER_NODES]
    europe_nodes = [n for n, r in node_region.items() if r in EUROPE_REGIONS]
    if not china_nodes or not europe_nodes: return df
    q_vals  = pd.to_numeric(df["QuantityTEU"], errors="coerce").dropna()
    q_min   = int(q_vals.min()) if len(q_vals) else 80
    q_max   = int(q_vals.max()) if len(q_vals) else 150
    lt_vals = pd.to_numeric(df["LT"], errors="coerce").dropna()
    lt_vals = lt_vals[lt_vals >= 300]
    lt_min, lt_max = (int(lt_vals.min()), int(lt_vals.max())) if len(lt_vals) else (360, 504)
    pen_col_exists = "PenaltyCost_per_TEU_h" in df.columns
    if pen_col_exists:
        pv = pd.to_numeric(df["PenaltyCost_per_TEU_h"], errors="coerce").dropna()
        pv = pv[(pv >= 1.0) & (pv <= 500.0)]
        pen_min, pen_max = (float(pv.min()), float(pv.max())) if len(pv) else (30.0, 100.0)
    else:
        pen_min, pen_max = 30.0, 100.0
    existing_ids = set(pd.to_numeric(df["BatchID"], errors="coerce").dropna().astype(int).tolist())
    next_id      = max(existing_ids) + 1 if existing_ids else 11
    rng          = np.random.default_rng(random_seed)
    new_rows     = []
    for i in range(20 - len(df)):
        new_rows.append({
            "BatchID": next_id + i,
            "OriginEN": str(rng.choice(china_nodes)),
            "DestEN":   str(rng.choice(europe_nodes)),
            "QuantityTEU": int(rng.integers(q_min, q_max + 1)),
            "ET": 0,
            "LT": int(rng.integers(lt_min, lt_max + 1)),
            "PenaltyCost_per_TEU_h": round(float(rng.uniform(pen_min, pen_max)), 2),
        })
    df_out = pd.concat([df, pd.DataFrame(new_rows)], ignore_index=True)
    print(f"[INFO] Batches augmented: {len(df)} -> {len(df_out)}")
    return df_out


# ════════════════════════════════════════════════════════
# Fixed mutation operator sampling  (adaptive roulette removed)
# ════════════════════════════════════════════════════════

_FIXED_OP_WEIGHTS = [W_ADD, W_DEL, W_MOD, W_MODE]
_FIXED_OP_TOTAL   = sum(_FIXED_OP_WEIGHTS)
_FIXED_OP_PROBS   = [w / _FIXED_OP_TOTAL for w in _FIXED_OP_WEIGHTS]


def sample_operator() -> str:
    r, cum = random.random(), 0.0
    for op, prob in zip(OPS, _FIXED_OP_PROBS):
        cum += prob
        if r <= cum: return op
    return OPS[-1]


def apply_mutation_op(ind, op, batch, path_lib, tt_dict, arc_lookup):
    if op == "add":  return mutate_add(ind, batch, path_lib)
    if op == "del":  return mutate_del(ind, batch)
    if op == "mod":  return mutate_mod(ind, batch)
    if op == "mode": return mutate_mode(ind, batch, tt_dict, arc_lookup)
    return False


def mutate_fixed(
    ind, batches, path_lib, tt_dict, arc_lookup,
    arcs,
    waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
    node_hold_cost=None, node_proc_cost=None,
    carbon_tax_map=None, trans_map=None, border_delay_map=None,
):
    batch = random.choice(batches)
    op    = sample_operator()
    ok    = apply_mutation_op(ind, op, batch, path_lib, tt_dict, arc_lookup)
    if not ok:
        return op, False
    repair_missing_allocations(ind, batches, path_lib)
    evaluate_individual(
        ind, batches, arcs, tt_dict,
        waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
        node_hold_cost=node_hold_cost, node_proc_cost=node_proc_cost,
        carbon_tax_map=carbon_tax_map, trans_map=trans_map,
        border_delay_map=border_delay_map)
    return op, True


# ════════════════════════════════════════════════════════
# Dominance (shared with SPEA2)
# ════════════════════════════════════════════════════════

def dominates(a, b):
    if a.feasible and not b.feasible: return True
    if b.feasible and not a.feasible: return False
    if a.feasible and b.feasible:
        return (all(x <= y for x, y in zip(a.objectives, b.objectives)) and
                any(x <  y for x, y in zip(a.objectives, b.objectives)))
    if a.penalty < b.penalty - 1e-12: return True
    if b.penalty < a.penalty - 1e-12: return False
    return (all(x <= y for x, y in zip(a.objectives, b.objectives)) and
            any(x <  y for x, y in zip(a.objectives, b.objectives)))


def unique_individuals_by_objectives(front, tol=1e-3):
    uniq, seen = [], []
    for ind in front:
        obj = ind.objectives
        if not any(all(abs(obj[i]-o[i]) <= tol for i in range(3)) for o in seen):
            seen.append(obj); uniq.append(ind)
    return uniq


# ════════════════════════════════════════════════════════
# Feasibility Boost  (identical to NSGA-II)
# ════════════════════════════════════════════════════════

def _select_boost_parents(population, topk=FEASIBLE_BOOST_TOPK_PARENTS):
    feasible   = sorted([i for i in population if i.feasible],
                        key=lambda x: sum(x.objectives))
    if len(feasible) >= topk: return feasible[:topk]
    infeasible = sorted([i for i in population if not i.feasible],
                        key=lambda x: x.penalty)
    return (feasible + infeasible)[:topk]


def feasibility_boost(
    population, batches, path_lib, tt_dict, arc_lookup,
    arcs,
    waiting_cost_per_teu_h, wait_emis_g_per_teu_h, eval_kwargs,
    boost_rounds=FEASIBLE_BOOST_ROUNDS,
    boost_mutation_rate=FEASIBLE_BOOST_MUTATION_RATE,
    topk_parents=FEASIBLE_BOOST_TOPK_PARENTS,
):
    parents  = _select_boost_parents(population, topk=topk_parents)
    new_inds = []
    if len(parents) < 2:
        for _ in range(boost_rounds):
            ind = greedy_initial_individual(batches, path_lib)
            repair_missing_allocations(ind, batches, path_lib)
            evaluate_individual(ind, batches, arcs, tt_dict,
                                waiting_cost_per_teu_h, wait_emis_g_per_teu_h, **eval_kwargs)
            new_inds.append(ind)
    else:
        for _ in range(boost_rounds):
            p1, p2 = random.sample(parents, 2)
            c1, c2 = crossover_hybrid(p1, p2, batches, tt_dict, arc_lookup)
            for child in (c1, c2):
                repair_missing_allocations(child, batches, path_lib)
                if random.random() < boost_mutation_rate:
                    mutate_fixed(
                        child, batches, path_lib, tt_dict, arc_lookup,
                        arcs,
                        waiting_cost_per_teu_h, wait_emis_g_per_teu_h, **eval_kwargs)
                new_inds.append(child)
            if len(new_inds) >= boost_rounds: break

    pop_id_to_idx   = {id(ind): idx for idx, ind in enumerate(population)}
    infeas_sorted   = sorted([i for i in population if not i.feasible],
                             key=lambda x: x.penalty, reverse=True)
    new_sorted      = sorted(new_inds, key=lambda x: (0 if x.feasible else 1, x.penalty))
    num_new_feas    = 0
    replaced_ids: set = set()
    for new_ind in new_sorted:
        for target in infeas_sorted:
            if id(target) in replaced_ids: continue
            if new_ind.feasible or new_ind.penalty < target.penalty:
                idx = pop_id_to_idx.get(id(target))
                if idx is not None:
                    population[idx] = new_ind
                    replaced_ids.add(id(target))
                    if new_ind.feasible: num_new_feas += 1
                break
    return population, num_new_feas


# ════════════════════════════════════════════════════════
# ★★★  SPEA2-specific: Fitness Assignment & Archive  ★★★
# ════════════════════════════════════════════════════════

def spea2_strength(combined: List[Individual]) -> Dict[int, int]:
    """S(i) = number of individuals that i dominates."""
    strength = {}
    for i, ind_i in enumerate(combined):
        s = 0
        for j, ind_j in enumerate(combined):
            if i != j and dominates(ind_i, ind_j):
                s += 1
        strength[id(ind_i)] = s
    return strength


def spea2_raw_fitness(combined: List[Individual],
                      strength: Dict[int, int]) -> Dict[int, float]:
    """R(i) = sum of S(j) for all j that dominate i."""
    raw = {}
    for i, ind_i in enumerate(combined):
        r = 0.0
        for j, ind_j in enumerate(combined):
            if i != j and dominates(ind_j, ind_i):
                r += strength[id(ind_j)]
        raw[id(ind_i)] = r
    return raw


def spea2_distances(combined: List[Individual]) -> np.ndarray:
    """Pairwise Euclidean distance matrix in normalised objective+penalty space."""
    objs = np.array([list(ind.objectives) for ind in combined], dtype=float)
    pens = np.array([ind.penalty for ind in combined], dtype=float)
    mat  = np.column_stack([objs, pens])
    mins = np.min(mat, axis=0)
    maxs = np.max(mat, axis=0)
    rng  = maxs - mins
    rng[rng < 1e-12] = 1.0
    mat_norm = (mat - mins) / rng
    n    = len(combined)
    dist = np.zeros((n, n), dtype=float)
    for i in range(n):
        diff    = mat_norm - mat_norm[i]
        dist[i] = np.sqrt(np.sum(diff ** 2, axis=1))
    return dist


def spea2_density(dist_matrix: np.ndarray, k: int) -> np.ndarray:
    """D(i) = 1 / (sigma_k(i) + 2)"""
    n       = dist_matrix.shape[0]
    density = np.zeros(n, dtype=float)
    for i in range(n):
        dists_i    = np.sort(dist_matrix[i])
        sigma_k    = dists_i[min(k, n - 1)] if n > 1 else 0.0
        density[i] = 1.0 / (sigma_k + 2.0)
    return density


def spea2_fitness_assignment(combined: List[Individual]) -> Dict[int, float]:
    """F(i) = R(i) + D(i)   (lower is better)."""
    n = len(combined)
    if n == 0: return {}
    strength = spea2_strength(combined)
    raw      = spea2_raw_fitness(combined, strength)
    dist_mat = spea2_distances(combined)
    k        = max(1, int(math.sqrt(n)))
    density  = spea2_density(dist_mat, k)
    fitness  = {}
    for i, ind in enumerate(combined):
        fitness[id(ind)] = raw[id(ind)] + density[i]
    return fitness


def spea2_environmental_selection(combined: List[Individual],
                                   archive_size: int) -> List[Individual]:
    """
    1. F(i) < 1  → into archive  (non-dominated individuals).
    2. If |archive| < archive_size: fill with best remaining by F(i).
    3. If |archive| > archive_size: truncate by nearest-neighbour removal.
    """
    fitness = spea2_fitness_assignment(combined)

    archive = [ind for ind in combined if fitness[id(ind)] < 1.0]

    if len(archive) < archive_size:
        remaining = sorted(
            [ind for ind in combined if fitness[id(ind)] >= 1.0],
            key=lambda x: fitness[id(x)])
        archive.extend(remaining[:archive_size - len(archive)])

    elif len(archive) > archive_size:
        while len(archive) > archive_size:
            dist_mat = spea2_distances(archive)
            n = len(archive)
            sorted_dists = []
            for i in range(n):
                d = np.sort(dist_mat[i])
                sorted_dists.append(d)
            remove_idx = 0
            for i in range(1, n):
                for ki in range(1, n):
                    if sorted_dists[i][ki] < sorted_dists[remove_idx][ki] - 1e-15:
                        remove_idx = i; break
                    elif sorted_dists[i][ki] > sorted_dists[remove_idx][ki] + 1e-15:
                        break
            archive.pop(remove_idx)

    return archive[:archive_size]


def spea2_tournament_select(archive: List[Individual],
                             fitness: Dict[int, float]) -> Individual:
    a, b = random.sample(archive, 2)
    return a if fitness.get(id(a), float("inf")) <= fitness.get(id(b), float("inf")) else b


# ════════════════════════════════════════════════════════
# HV calculator (Monte Carlo)
# ════════════════════════════════════════════════════════

class HypervolumeCalculator:
    def __init__(self, ref_point, num_samples=2000, seed=None):
        self.ref_point  = np.array(ref_point, dtype=float)
        self.num_samples= int(num_samples)
        rng = np.random.default_rng(seed)
        self.samples    = rng.uniform(low=0.0, high=self.ref_point,
                                      size=(self.num_samples, 3))

    def calculate_points(self, points):
        if not points: return 0.0
        front = np.array(points, dtype=float)
        valid = front[np.all(front <= self.ref_point, axis=1)]
        if len(valid) == 0: return 0.0
        S   = self.samples[:, np.newaxis, :]
        O   = valid[np.newaxis, :, :]
        dom = np.any(np.all(O <= S, axis=2), axis=1)
        return float(np.sum(dom) / self.num_samples)


# ════════════════════════════════════════════════════════
# Metrics: P*, IGD+, Spacing
# ════════════════════════════════════════════════════════

def dominates_obj(a, b):
    return all(a[i] <= b[i] for i in range(3)) and any(a[i] < b[i] for i in range(3))


def nondominated_set(points):
    pts = unique_objective_tuples(points, tol=1e-9)
    return [p for i, p in enumerate(pts)
            if not any(dominates_obj(q, p) for j, q in enumerate(pts) if i != j)]


def normalize_points(points, mins, maxs):
    out = []
    for p in points:
        pp = []
        for i in range(3):
            rng = maxs[i] - mins[i]
            pp.append(0.0 if rng <= 1e-12 else (p[i]-mins[i])/rng)
        out.append(tuple(pp))
    return out


def clip_points(points, ref):
    return [tuple(min(max(p[i], 0.0), ref[i]) for i in range(3)) for p in points]


def igd_plus(P_star, A):
    if not P_star or not A: return float("inf")
    P, Q = np.array(P_star, dtype=float), np.array(A, dtype=float)
    return float(np.mean([
        float(np.min(np.sqrt(np.sum(np.maximum(Q - p, 0.0)**2, axis=1))))
        for p in P]))


def spacing_metric(A):
    if not A or len(A) < 2: return 0.0
    Q = np.array(A, dtype=float)
    n = Q.shape[0]
    dmin = []
    for i in range(n):
        diff = Q - Q[i]
        d    = np.sqrt(np.sum(diff**2, axis=1))
        d[i] = np.inf
        dmin.append(float(np.min(d)))
    dmin = np.array(dmin)
    return float(np.sqrt(np.sum((dmin - np.mean(dmin))**2) / max(1, n-1)))


def build_P_star_fast(run_front_hist, tail_gens=PSTAR_TAIL_GENS,
                      cap_per_gen=PSTAR_CAP_PER_GEN, max_total=PSTAR_MAX_TOTAL):
    pts = []
    for hist in run_front_hist:
        tail = hist[-tail_gens:] if tail_gens > 0 else hist
        for gen_front in tail:
            pts.extend(gen_front[:cap_per_gen])
            if len(pts) >= max_total: break
        if len(pts) >= max_total: break
    return nondominated_set(pts)


# ════════════════════════════════════════════════════════
# Metrics export & plotting helpers
# ════════════════════════════════════════════════════════

def export_metrics_csv(
    gen_arr, hv_mean, hv_std, igd_mean, igd_std,
    sp_mean, sp_std, fr_mean, fr_std, frs_mean, frs_std,
    min_cost_best, min_emis_best, min_time_best,
    boost_hist_best, vio_mean_dict_mean,
    out_csv="metrics_per_generation.csv",
):
    df = pd.DataFrame({
        "generation":           gen_arr,
        "HV_norm_mean":         hv_mean,   "HV_norm_std":         hv_std,
        "IGD_plus_mean":        igd_mean,  "IGD_plus_std":        igd_std,
        "Spacing_mean":         sp_mean,   "Spacing_std":         sp_std,
        "FeasRatio_soft_mean":  fr_mean,   "FeasRatio_soft_std":  fr_std,
        "FeasRatio_strict_mean":frs_mean,  "FeasRatio_strict_std":frs_std,
        "MinCost_bestrun":      min_cost_best,
        "MinEmission_gCO2_bestrun": min_emis_best,
        "MinTime_h_bestrun":    min_time_best,
        "boost_triggered_bestrun": boost_hist_best if boost_hist_best else [0]*len(gen_arr),
    })
    for k, series in vio_mean_dict_mean.items():
        df[f"vio_{k}_mean"] = series
    df.to_csv(out_csv, index=False)
    print(f"[EXPORT] Per-generation metrics → {out_csv}")
    return df


def export_run_summary_excel(run_rows, hv_runs, igd_runs, sp_runs,
                              out_xlsx="run_summary.xlsx"):
    df = pd.DataFrame(run_rows)
    df["final_HV_norm"]  = hv_runs[:, -1]
    df["final_IGD_plus"] = np.array(igd_runs)[:, -1]
    df["final_Spacing"]  = np.array(sp_runs)[:, -1]
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="RunSummary", index=False)
        ws = writer.sheets["RunSummary"]
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            fill = PatternFill("solid", fgColor="1F4E79")
            font = Font(color="FFFFFF", bold=True)
            for ci, col in enumerate(df.columns, 1):
                c = ws.cell(row=1, column=ci)
                c.fill, c.font = fill, font
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(ci)].width = min(len(col)+4, 22)
            ws.freeze_panes = "A2"
        except Exception:
            pass
    print(f"[EXPORT] Run summary → {out_xlsx}")
    return df


def save_pareto_solutions(pareto, batches, filename="result.txt"):
    pareto = unique_individuals_by_objectives([i for i in pareto if i.feasible], tol=1e-3)
    with open(filename, "w", encoding="utf-8") as f:
        f.write("===== SPEA2 Pareto Solutions (Feasible, No Node Capacity) =====\n\n")
        if not pareto:
            f.write("NO FEASIBLE SOLUTION FOUND.\n"); return
        for i, ind in enumerate(pareto):
            c, e, t = ind.objectives
            f.write(f"===== Pareto Sol {i} =====\n")
            f.write(f"Cost={c:.6f}  Emission_gCO2={e:.6f}  Time={t:.6f}\n")
            f.write(f"Penalty={ind.penalty:.6f}  Feasible={ind.feasible}\n")
            f.write(f"Breakdown={ind.vio_breakdown}\n\n")
            for b in batches:
                key   = (b.origin, b.destination, b.batch_id)
                allocs= ind.od_allocations.get(key, [])
                if not allocs: continue
                f.write(f"Batch {b.batch_id}: {b.origin} -> {b.destination}, Q={b.quantity}\n")
                for a in allocs: f.write(str(a) + "\n")
                f.write("\n")
            f.write("\n")
    print(f"[EXPORT] {len(pareto)} feasible Pareto solutions → {filename}")


def export_pareto_points_json(pareto, batches, out_json="pareto_points.json"):
    out = []
    for ind in pareto:
        sol = {
            "objectives": {
                "cost":         float(ind.objectives[0]),
                "emission_gCO2":float(ind.objectives[1]),
                "time_h":       float(ind.objectives[2]),
                "penalty":      float(ind.penalty),
            },
            "feasible":      bool(ind.feasible),
            "vio_breakdown": {k: float(v) for k, v in (ind.vio_breakdown or {}).items()},
            "allocations":   []
        }
        for b in batches:
            key = (b.origin, b.destination, b.batch_id)
            blk = {"batch_id": int(b.batch_id), "origin": b.origin,
                   "destination": b.destination, "paths": []}
            for a in ind.od_allocations.get(key, []):
                blk["paths"].append({"share": float(a.share),
                                     "nodes": list(a.path.nodes),
                                     "modes": list(a.path.modes)})
            sol["allocations"].append(blk)
        out.append(sol)
    with open(out_json, "w", encoding="utf-8") as f:
        _json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"[EXPORT] Pareto JSON → {out_json}")


# ════════════════════════════════════════════════════════
# Plotting functions  (identical to NSGA-II)
# ════════════════════════════════════════════════════════

def plot_hv_curve(gen, hv_mean, hv_std, save="plot_HV.png"):
    fig, ax = plt.subplots(figsize=(9, 4), dpi=180)
    ax.plot(gen, hv_mean, lw=2.2, color="#9C27B0", label="HV (normalised, mean)")
    ax.fill_between(gen, hv_mean-hv_std, hv_mean+hv_std,
                    alpha=0.22, color="#9C27B0", label="±std")
    ax.set_xlabel("Generation"); ax.set_ylabel("HV (normalised)")
    ax.set_ylim(bottom=0)
    ax.set_title("SPEA2: Hypervolume — Normalised (mean ± std over runs)")
    ax.grid(True, ls=":", alpha=0.5); ax.legend()
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_igd_curve(gen, igd_mean, igd_std, save="plot_IGDplus.png"):
    ip = np.where(np.isfinite(igd_mean), igd_mean, np.nan)
    is_= np.where(np.isfinite(igd_std),  igd_std,  np.nan)
    fig, ax = plt.subplots(figsize=(9, 4), dpi=180)
    ax.plot(gen, ip, lw=2.2, color="#E53935", label="IGD+ (mean)")
    ax.fill_between(gen, np.maximum(ip-is_, 0), ip+is_, alpha=0.2, color="#E53935", label="±std")
    ax.set_xlabel("Generation"); ax.set_ylabel("IGD+ (lower is better)")
    ax.set_title("SPEA2: IGD+ (mean ± std over runs)")
    ax.grid(True, ls=":", alpha=0.5); ax.legend()
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_spacing_curve(gen, sp_mean, sp_std, save="plot_Spacing.png"):
    fig, ax = plt.subplots(figsize=(9, 4), dpi=180)
    ax.plot(gen, sp_mean, lw=2.2, color="#00897B", label="Spacing (mean)")
    ax.fill_between(gen, np.maximum(sp_mean-sp_std, 0), sp_mean+sp_std,
                    alpha=0.2, color="#00897B", label="±std")
    ax.set_xlabel("Generation"); ax.set_ylabel("Spacing (lower is better)")
    ax.set_title("SPEA2: Spacing — Diversity Metric (mean ± std over runs)")
    ax.grid(True, ls=":", alpha=0.5); ax.legend()
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_feasible_ratio_curve(gen, fr_mean, fr_std, frs_mean, frs_std,
                               boost_hist=None, save="plot_FeasibleRatio.png"):
    fig, ax = plt.subplots(figsize=(9, 4), dpi=180)
    ax.plot(gen, fr_mean, lw=2.2, color="#1976D2", label="Soft (mean)")
    ax.fill_between(gen, np.clip(fr_mean-fr_std, 0, 1), np.clip(fr_mean+fr_std, 0, 1),
                    alpha=0.18, color="#1976D2", label="soft ±std")
    ax.plot(gen, frs_mean, lw=2.0, ls="--", color="#FB8C00", label="Strict (mean)")
    ax.fill_between(gen, np.clip(frs_mean-frs_std, 0, 1), np.clip(frs_mean+frs_std, 0, 1),
                    alpha=0.13, color="#FB8C00", label="strict ±std")
    if boost_hist:
        bgs = [g for g, v in enumerate(boost_hist) if v > 0]
        if bgs:
            ax.scatter(bgs, [fr_mean[g] for g in bgs if g < len(fr_mean)],
                       marker="^", color="red", s=40, zorder=5, label=f"Boost ({len(bgs)} gens)")
    ax.set_ylim(-0.02, 1.05)
    ax.set_xlabel("Generation"); ax.set_ylabel("Feasible Ratio")
    ax.set_title(f"SPEA2: Feasible Ratio (threshold={MIN_FEASIBLE_SOLUTIONS})")
    ax.grid(True, ls=":", alpha=0.5); ax.legend(fontsize=8)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_min_objectives(gen, min_cost, min_emis, min_time, save="plot_MinObjectives.png"):
    fig, axes = plt.subplots(3, 1, figsize=(10, 9), dpi=180, sharex=True)
    fig.subplots_adjust(hspace=0.35)
    pairs = [(min_cost, "Min Cost ($)", "#E53935"),
             (min_emis, "Min Emission (gCO₂)", "#00897B"),
             (min_time, "Min Time (h)", "#FB8C00")]
    for ax, (data, ylabel, color) in zip(axes, pairs):
        ax.plot(gen, data, lw=2.2, color=color)
        mask = np.isfinite(data)
        if mask.any():
            fg = int(np.where(mask)[0][0])
            ax.axvline(fg, color="green", ls="--", alpha=0.7, lw=1.0,
                       label=f"First feasible gen {fg}")
            ax.legend(fontsize=8)
        ax.set_ylabel(ylabel, fontsize=9)
        ax.grid(True, ls=":", alpha=0.5)
    axes[-1].set_xlabel("Generation")
    fig.suptitle("SPEA2: Best-Run Min Objectives per Generation (feasible only)",
                 fontsize=12, fontweight="bold")
    plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_convergence_combined(gen, hv_mean, hv_std, igd_mean, igd_std,
                               save="plot_Convergence.png"):
    ip = np.where(np.isfinite(igd_mean), igd_mean, np.nan)
    is_= np.where(np.isfinite(igd_std),  igd_std,  np.nan)
    fig, ax1 = plt.subplots(figsize=(10, 4), dpi=180)
    ax1.plot(gen, hv_mean, lw=2.0, color="#9C27B0", label="HV (mean)")
    ax1.fill_between(gen, hv_mean-hv_std, hv_mean+hv_std, alpha=0.18, color="#9C27B0")
    ax1.set_xlabel("Generation"); ax1.set_ylabel("HV (normalised)", color="#9C27B0")
    ax1.tick_params(axis="y", labelcolor="#9C27B0"); ax1.set_ylim(bottom=0)
    ax2 = ax1.twinx()
    ax2.plot(gen, ip, lw=2.0, ls="--", color="#E53935", label="IGD+ (mean)")
    ax2.fill_between(gen,
                     np.where(np.isfinite(ip-is_), ip-is_, np.nan),
                     np.where(np.isfinite(ip+is_), ip+is_, np.nan),
                     alpha=0.13, color="#E53935")
    ax2.set_ylabel("IGD+ (lower is better)", color="#E53935")
    ax2.tick_params(axis="y", labelcolor="#E53935")
    lines  = ax1.get_legend_handles_labels()[0] + ax2.get_legend_handles_labels()[0]
    labels = ax1.get_legend_handles_labels()[1] + ax2.get_legend_handles_labels()[1]
    ax1.legend(lines, labels, loc="lower right", fontsize=8)
    ax1.grid(True, ls=":", alpha=0.5)
    plt.title("SPEA2: Convergence — HV & IGD+ (mean ± std over runs)")
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_pareto_3d(pareto_points, save="plot_Pareto3D.png", title="SPEA2 Pareto Front"):
    if not pareto_points: return
    A = _finite_points_array(pareto_points)
    if A.shape[0] == 0: return
    fig = plt.figure(figsize=(7, 6), dpi=200)
    ax  = fig.add_subplot(111, projection="3d")
    sc  = ax.scatter(A[:, 0], A[:, 1], A[:, 2], c=A[:, 0], cmap="viridis", s=40, alpha=0.9)
    plt.colorbar(sc, ax=ax, pad=0.1, fraction=0.04, label="Cost ($)")
    ax.set_xlabel("Cost ($)"); ax.set_ylabel("Emission (gCO₂)"); ax.set_zlabel("Time (h)")
    ax.set_title(title); ax.grid(True, ls=":", alpha=0.4)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_pareto_2d_projections(pareto_points, save="plot_Pareto2D.png"):
    if not pareto_points: return
    A = _finite_points_array(pareto_points)
    if A.shape[0] == 0: return
    fig, axes = plt.subplots(1, 3, figsize=(14, 4.5), dpi=180)
    pairs = [(0,1,"Cost ($)","Emission (gCO₂)"),
             (0,2,"Cost ($)","Time (h)"),
             (1,2,"Emission (gCO₂)","Time (h)")]
    for ax, (xi, yi, xl, yl) in zip(axes, pairs):
        ax.scatter(A[:, xi], A[:, yi], s=30, alpha=0.85, c=A[:, 0], cmap="viridis")
        ax.set_xlabel(xl); ax.set_ylabel(yl)
        ax.set_title(f"{xl} vs {yl}")
        ax.grid(True, ls=":", alpha=0.45)
    fig.suptitle("SPEA2: Final Pareto Front — 2D Projections (Best Run)",
                 fontsize=11, fontweight="bold")
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_mutation_adaptive_prob(gen, prob_mean, save="plot_MutationProb.png"):
    fig, ax = plt.subplots(figsize=(10, 4), dpi=180)
    ax.stackplot(gen, prob_mean[0], prob_mean[1], prob_mean[2], prob_mean[3],
                 labels=["Add","Delete","Modify","Mode"], alpha=0.82)
    ax.set_ylim(0, 1); ax.set_xlabel("Generation"); ax.set_ylabel("Selection Probability")
    ax.set_title("SPEA2: Adaptive Roulette Operator Probability (mean over runs)")
    ax.legend(loc="upper right"); ax.grid(axis="y", ls="--", alpha=0.3)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_runtime_per_run(run_rows, save="plot_Runtime.png"):
    df  = pd.DataFrame(run_rows)
    avg = df["runtime_s"].mean()
    colors = ["#E53935" if r > avg else "#1976D2" for r in df["runtime_s"]]
    fig, ax = plt.subplots(figsize=(max(8, len(df)*0.35), 4), dpi=180)
    ax.bar(df["run_id"], df["runtime_s"], color=colors, alpha=0.85)
    ax.axhline(avg, color="black", ls="--", lw=1.5, label=f"Mean = {avg:.1f}s")
    ax.set_xlabel("Run ID"); ax.set_ylabel("Runtime (s)")
    ax.set_title("SPEA2: Runtime per Run")
    ax.legend(); ax.grid(axis="y", ls=":", alpha=0.5)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_summary_metrics_table(hv_runs, igd_runs, sp_runs, fr_runs, frs_runs,
                                run_rows, save="plot_SummaryTable.png"):
    metrics = {
        "HV_norm (↑)":     hv_runs[:, -1],
        "IGD+ (↓)":        np.array(igd_runs)[:, -1],
        "Spacing (↓)":     np.array(sp_runs)[:, -1],
        "FeasRatio soft (↑)": fr_runs[:, -1],
        "FeasRatio strict (↑)": frs_runs[:, -1],
        "Runtime (s)": np.array([r["runtime_s"] for r in run_rows]),
    }
    table_data = []
    for name, vals in metrics.items():
        finite = vals[np.isfinite(vals)]
        if len(finite) == 0:
            table_data.append([name, "—", "—", "—", "—"])
        else:
            table_data.append([name,
                                f"{np.min(finite):.4f}", f"{np.max(finite):.4f}",
                                f"{np.mean(finite):.4f}", f"{np.std(finite):.4f}"])
    fig, ax = plt.subplots(figsize=(10, 3), dpi=150)
    ax.axis("off")
    tbl = ax.table(cellText=table_data,
                   colLabels=["Metric","Min","Max","Mean","Std"],
                   cellLoc="center", loc="center", bbox=[0.0, 0.0, 1.0, 1.0])
    tbl.auto_set_font_size(False); tbl.set_fontsize(10)
    for j in range(5):
        tbl[(0, j)].set_facecolor("#1F4E79")
        tbl[(0, j)].set_text_props(color="white", fontweight="bold")
    for i in range(len(table_data)):
        clr = "#EBF3FB" if i % 2 == 0 else "#FFFFFF"
        for j in range(5):
            tbl[(i+1, j)].set_facecolor(clr)
    fig.suptitle(f"SPEA2: Summary Statistics over {len(run_rows)} Runs (Final Generation)",
                 fontsize=11, fontweight="bold", y=1.02)
    plt.tight_layout(); plt.savefig(save, bbox_inches="tight"); plt.close()
    print(f"[PLOT] {save}")


def plot_objectives_scatter_by_run(all_pareto_pts_by_run, save="plot_ParetoByRun.png"):
    fig, ax = plt.subplots(figsize=(9, 5), dpi=180)
    cmap = plt.cm.tab20
    for r_idx, pts in enumerate(all_pareto_pts_by_run):
        arr = _finite_points_array(pts)
        if arr.shape[0] == 0: continue
        color = cmap(r_idx / max(len(all_pareto_pts_by_run), 1))
        ax.scatter(arr[:, 0], arr[:, 2], s=18, alpha=0.6, color=color, label=f"Run {r_idx}")
    ax.set_xlabel("Cost ($)"); ax.set_ylabel("Time (h)")
    ax.set_title("SPEA2: Final Pareto Front — Cost vs Time (all runs)")
    ax.grid(True, ls=":", alpha=0.5)
    if len(all_pareto_pts_by_run) <= 12:
        ax.legend(fontsize=7, ncol=3)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_min_cost_trend(gen, min_cost_mean, min_cost_std=None,
                         save="plot_MinCost_trend.png"):
    fig, ax = plt.subplots(figsize=(10, 4), dpi=180)
    ax.plot(gen, min_cost_mean, lw=2.2, color="#E53935", label="Min Cost (mean)")
    if min_cost_std is not None:
        ax.fill_between(gen,
            np.where(np.isfinite(min_cost_mean-min_cost_std), min_cost_mean-min_cost_std, np.nan),
            np.where(np.isfinite(min_cost_mean+min_cost_std), min_cost_mean+min_cost_std, np.nan),
            alpha=0.20, color="#E53935", label="±std")
    mask = np.isfinite(min_cost_mean)
    if mask.any():
        fg = int(np.where(mask)[0][0])
        ax.axvline(fg, color="green", ls="--", lw=1.2, alpha=0.7, label=f"First feasible gen {fg}")
    ax.set_xlabel("Generation"); ax.set_ylabel("Min Cost ($)")
    ax.set_title("SPEA2: Min Transport Cost per Generation (mean ± std)")
    ax.grid(True, ls=":", alpha=0.5); ax.legend(fontsize=9)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_min_emission_trend(gen, min_emis_mean, min_emis_std=None,
                             save="plot_MinEmission_trend.png"):
    fig, ax = plt.subplots(figsize=(10, 4), dpi=180)
    ax.plot(gen, min_emis_mean, lw=2.2, color="#00897B", label="Min Emission (mean)")
    if min_emis_std is not None:
        ax.fill_between(gen,
            np.where(np.isfinite(min_emis_mean-min_emis_std), min_emis_mean-min_emis_std, np.nan),
            np.where(np.isfinite(min_emis_mean+min_emis_std), min_emis_mean+min_emis_std, np.nan),
            alpha=0.20, color="#00897B", label="±std")
    mask = np.isfinite(min_emis_mean)
    if mask.any():
        fg = int(np.where(mask)[0][0])
        ax.axvline(fg, color="green", ls="--", lw=1.2, alpha=0.7, label=f"First feasible gen {fg}")
    ax.set_xlabel("Generation"); ax.set_ylabel("Min Emission (gCO₂)")
    ax.set_title("SPEA2: Min Emission per Generation (mean ± std)")
    ax.grid(True, ls=":", alpha=0.5); ax.legend(fontsize=9)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_min_time_trend(gen, min_time_mean, min_time_std=None,
                         save="plot_MinTime_trend.png"):
    fig, ax = plt.subplots(figsize=(10, 4), dpi=180)
    ax.plot(gen, min_time_mean, lw=2.2, color="#FB8C00", label="Min Time (mean)")
    if min_time_std is not None:
        ax.fill_between(gen,
            np.where(np.isfinite(min_time_mean-min_time_std), min_time_mean-min_time_std, np.nan),
            np.where(np.isfinite(min_time_mean+min_time_std), min_time_mean+min_time_std, np.nan),
            alpha=0.20, color="#FB8C00", label="±std")
    mask = np.isfinite(min_time_mean)
    if mask.any():
        fg = int(np.where(mask)[0][0])
        ax.axvline(fg, color="green", ls="--", lw=1.2, alpha=0.7, label=f"First feasible gen {fg}")
    ax.set_xlabel("Generation"); ax.set_ylabel("Min Time (h)")
    ax.set_title("SPEA2: Min Makespan per Generation (mean ± std)")
    ax.grid(True, ls=":", alpha=0.5); ax.legend(fontsize=9)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_pareto_size_trend(gen, pareto_size_mean, pareto_size_std=None,
                            pareto_size_best=None, save="plot_ParetoSize_trend.png"):
    fig, ax = plt.subplots(figsize=(10, 4), dpi=180)
    ax.plot(gen, pareto_size_mean, lw=2.2, color="#5C6BC0", label="Pareto F0 size (mean)")
    if pareto_size_std is not None:
        ax.fill_between(gen,
            np.maximum(pareto_size_mean-pareto_size_std, 0),
            pareto_size_mean+pareto_size_std,
            alpha=0.20, color="#5C6BC0", label="±std")
    if pareto_size_best is not None:
        ax.plot(gen, pareto_size_best, lw=1.5, ls="--", color="#E91E63",
                alpha=0.85, label="Best run")
    ax.set_xlabel("Generation")
    ax.set_ylabel("Number of solutions in Pareto Front 0")
    ax.set_title("SPEA2: Pareto Front Size Evolution")
    ax.grid(True, ls=":", alpha=0.5); ax.legend(fontsize=9)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_total_solution_count(gen, total_feas_mean, total_feas_std=None,
                               total_feas_best=None, save="plot_SolutionCount_trend.png"):
    fig, ax = plt.subplots(figsize=(10, 4), dpi=180)
    ax.plot(gen, total_feas_mean, lw=2.2, color="#1976D2", label="Feasible solutions (mean)")
    if total_feas_std is not None:
        ax.fill_between(gen,
            np.maximum(total_feas_mean-total_feas_std, 0),
            total_feas_mean+total_feas_std,
            alpha=0.20, color="#1976D2", label="±std")
    if total_feas_best is not None:
        ax.plot(gen, total_feas_best, lw=1.5, ls="--", color="#E91E63",
                alpha=0.85, label="Best run")
    ax.set_xlabel("Generation")
    ax.set_ylabel("Number of feasible solutions")
    ax.set_title("SPEA2: Feasible Solution Count Evolution")
    ax.grid(True, ls=":", alpha=0.5); ax.legend(fontsize=9)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_pareto_2d_allruns(all_pareto_pts_by_run, best_run_idx=None,
                            save="plot_Pareto2D_allruns.png"):
    fig, axes = plt.subplots(1, 3, figsize=(16, 5), dpi=180)
    pairs = [(0,1,"Cost ($)","Emission (gCO₂)"),
             (0,2,"Cost ($)","Time (h)"),
             (1,2,"Emission (gCO₂)","Time (h)")]
    for ax, (xi, yi, xl, yl) in zip(axes, pairs):
        for r_idx, pts in enumerate(all_pareto_pts_by_run):
            arr = _finite_points_array(pts)
            if arr.shape[0] == 0: continue
            if best_run_idx is not None and r_idx == best_run_idx: continue
            ax.scatter(arr[:, xi], arr[:, yi], s=12, alpha=0.35, color="#90CAF9", zorder=2)
        if best_run_idx is not None:
            ab = _finite_points_array(all_pareto_pts_by_run[best_run_idx])
            if ab.shape[0] > 0:
                sc = ax.scatter(ab[:, xi], ab[:, yi], s=30, alpha=0.95,
                                c=ab[:, 0], cmap="viridis", zorder=5,
                                label=f"Best run #{best_run_idx}")
                plt.colorbar(sc, ax=ax, fraction=0.04, pad=0.02, label="Cost ($)")
                ax.legend(fontsize=8)
        ax.set_xlabel(xl); ax.set_ylabel(yl)
        ax.set_title(f"{xl} vs {yl}")
        ax.grid(True, ls=":", alpha=0.45)
    fig.suptitle("SPEA2: Pareto 2D — All Runs (grey) + Best Run (colour)",
                 fontsize=11, fontweight="bold")
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


def plot_pareto_3d_allruns(all_pareto_pts_by_run, best_run_idx=None,
                            save="plot_Pareto3D_allruns.png"):
    fig = plt.figure(figsize=(9, 7), dpi=200)
    ax  = fig.add_subplot(111, projection="3d")
    for r_idx, pts in enumerate(all_pareto_pts_by_run):
        arr = _finite_points_array(pts)
        if arr.shape[0] == 0: continue
        if best_run_idx is not None and r_idx == best_run_idx: continue
        ax.scatter(arr[:, 0], arr[:, 1], arr[:, 2], s=8, alpha=0.25, color="#BDBDBD", zorder=2)
    if best_run_idx is not None:
        ab = _finite_points_array(all_pareto_pts_by_run[best_run_idx])
        if ab.shape[0] > 0:
            sc = ax.scatter(ab[:, 0], ab[:, 1], ab[:, 2],
                            c=ab[:, 0], cmap="viridis", s=45, alpha=0.95, zorder=5,
                            label=f"Best run #{best_run_idx}")
            plt.colorbar(sc, ax=ax, pad=0.1, fraction=0.035, label="Cost ($)")
            ax.legend(fontsize=9)
    ax.set_xlabel("Cost ($)"); ax.set_ylabel("Emission (gCO₂)"); ax.set_zlabel("Time (h)")
    ax.set_title("SPEA2: 3D Pareto — All Runs (grey) + Best Run (colour)")
    ax.grid(True, ls=":", alpha=0.4)
    plt.tight_layout(); plt.savefig(save); plt.close()
    print(f"[PLOT] {save}")


# ════════════════════════════════════════════════════════
# Extract helpers
# ════════════════════════════════════════════════════════

def extract_pareto_size_per_gen(run_front_hist):
    runs = len(run_front_hist)
    if runs == 0: return np.array([])
    generations = len(run_front_hist[0])
    mat = np.zeros((runs, generations), dtype=float)
    for r, hist in enumerate(run_front_hist):
        for g, gf in enumerate(hist):
            mat[r, g] = float(len(gf))
    return mat


def extract_min_obj_per_gen_allruns(run_front_hist, obj_idx):
    runs = len(run_front_hist)
    if runs == 0: return np.array([]), np.array([])
    generations = len(run_front_hist[0])
    mat = np.full((runs, generations), np.nan)
    for r, hist in enumerate(run_front_hist):
        for g, gf in enumerate(hist):
            arr = _finite_points_array(gf)
            if arr.shape[0] > 0:
                mat[r, g] = float(np.min(arr[:, obj_idx]))
    for r in range(runs):
        mat[r] = _ffill_nan(mat[r])
    return np.nanmean(mat, axis=0), np.nanstd(mat, axis=0)


# ════════════════════════════════════════════════════════
# Batch preparation: sample K batches, scale α
# ════════════════════════════════════════════════════════

def prepare_batches(raw_batches: List[Batch], K: int, alpha: float,
                    node_region: dict, random_seed: int = 2026) -> List[Batch]:
    rng = np.random.default_rng(random_seed)
    if len(raw_batches) >= K:
        idx      = rng.choice(len(raw_batches), size=K, replace=False)
        selected = [deepcopy(raw_batches[int(i)]) for i in idx]
    else:
        selected  = deepcopy(raw_batches)
        extra_idx = rng.choice(len(raw_batches),
                               size=K - len(raw_batches), replace=True)
        next_id   = max(b.batch_id for b in raw_batches) + 1
        for i, ei in enumerate(extra_idx):
            bc           = deepcopy(raw_batches[int(ei)])
            bc.batch_id  = next_id + i
            selected.append(bc)
    for b in selected:
        window = b.LT - b.ET
        b.LT   = b.ET + max(alpha * window, 1.0)
    for new_id, b in enumerate(selected):
        b.batch_id = new_id
    return selected


# ════════════════════════════════════════════════════════
# ★★★  Core SPEA2 runner  ★★★
# ════════════════════════════════════════════════════════

def run_spea2_analytics(filename="data.xlsx", pop_size=125,
                        generations=400, _preloaded=None):
    """
    SPEA2 main loop.
    _preloaded: same tuple format as NSGA-II version.
    """
    if _preloaded is not None:
        (node_names, node_region, node_hold_cost, node_proc_cost,
         arcs, timetables, batches,
         waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
         carbon_tax_map, emission_factor_map, mode_speeds_map,
         trans_map, border_delay_map, path_lib) = _preloaded
        tt_dict    = build_timetable_dict(timetables)
        arc_lookup = build_arc_lookup(arcs)
    else:
        print("Loading data...")
        (node_names, node_region, node_hold_cost, node_proc_cost,
         arcs, timetables, batches,
         waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
         carbon_tax_map, emission_factor_map, mode_speeds_map,
         trans_map, border_delay_map) = load_network_from_extended(filename)
        tt_dict    = build_timetable_dict(timetables)
        arc_lookup = build_arc_lookup(arcs)
        print("Building path library...")
        path_lib = build_path_library(
            node_names, node_region, arcs, batches, tt_dict, arc_lookup)
        sanity_check_path_lib(batches, path_lib)

    archive_size = int(pop_size * ARCHIVE_SIZE_RATIO)

    eval_kwargs = dict(
        node_hold_cost=node_hold_cost,   node_proc_cost=node_proc_cost,
        carbon_tax_map=carbon_tax_map,   trans_map=trans_map,
        border_delay_map=border_delay_map,
    )

    # ── Initialise population ─────────────────────────────
    population: List[Individual] = []
    n_greedy = max(1, pop_size // 3)
    for i in range(pop_size):
        ind = (greedy_initial_individual(batches, path_lib) if i < n_greedy
               else random_initial_individual(batches, path_lib))
        repair_missing_allocations(ind, batches, path_lib)
        evaluate_individual(ind, batches, arcs, tt_dict,
                            waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
                            **eval_kwargs)
        population.append(ind)

    archive: List[Individual] = []

    # ── History tracking ──────────────────────────────────
    front_hist_objs: List[List[Tuple]]     = []
    feasible_ratio_hist:        List[float] = []
    feasible_ratio_strict_hist: List[float] = []
    vio_mean_hist = {k: [] for k in ["miss_alloc","miss_tt","cap_excess",
                                     "late_teu_h","wait_teu_h"]}
    mut_tracker = {
        "attempt": {op: [0]   * generations for op in OPS},
        "success": {op: [0]   * generations for op in OPS},
        "prob":    {op: [_FIXED_OP_PROBS[i]] * generations for i, op in enumerate(OPS)},
    }
    boost_trigger_hist:      List[int] = []
    boost_new_feasible_hist: List[int] = []

    _run_start = time.perf_counter()
    _prev_best = [float("inf")] * 3

    # ═════════════════════════════════════════════════════
    #  SPEA2 main generational loop
    # ═════════════════════════════════════════════════════
    for gen in range(generations):

        # ── Step 1: Fitness assignment on P ∪ A ──────────
        combined = population + archive
        fitness  = spea2_fitness_assignment(combined)

        # ── Step 2: Environmental selection → new archive ─
        archive = spea2_environmental_selection(combined, archive_size)

        # ── Record metrics from archive ──────────────────
        feas_archive = [ind for ind in archive if ind.feasible]
        display      = unique_individuals_by_objectives(
            feas_archive if feas_archive else archive, tol=1e-3)

        front_hist_objs.append(
            [ind.objectives for ind in display if ind.feasible])

        feasible_ratio_hist.append(
            sum(1 for i in archive if i.feasible) / max(len(archive), 1))
        feasible_ratio_strict_hist.append(
            sum(1 for i in archive if i.feasible_hard) / max(len(archive), 1))

        for k in vio_mean_hist:
            vals = [ind.vio_breakdown.get(k, 0.0) for ind in archive]
            vio_mean_hist[k].append(float(np.mean(vals)) if vals else 0.0)

        # ── Logging ──────────────────────────────────────
        feas_total = sum(1 for i in archive if i.feasible)
        elapsed    = time.perf_counter() - _run_start
        feas_inds  = [i for i in archive if i.feasible]
        if feas_inds:
            cur    = [min(i.objectives[j] for i in feas_inds) for j in range(3)]
            d      = ["↓" if cur[j] < _prev_best[j] - 1e-3 else "→" for j in range(3)]
            _prev_best = cur
            obj_str = (f"Cost={cur[0]:.3e}{d[0]} "
                       f"Emis={cur[1]:.3e}{d[1]} "
                       f"Time={cur[2]:.1f}h{d[2]}")
        else:
            obj_str = "No feasible solutions yet"

        best_pen = min(i.penalty for i in archive) if archive else float("inf")
        sep      = "=" * 72
        print(f"\n{sep}")
        print(f"  [SPEA2] Gen {gen:03d}/{generations-1}  |  {elapsed:.1f}s elapsed")
        print(f"  Archive feasible: {feas_total}/{len(archive)} "
              f"({feasible_ratio_hist[-1]:.1%})"
              f"  |  Display={len(display)}"
              f"  |  BestPenalty={best_pen:.2e}")
        print(f"  Best feasible: {obj_str}")
        print(sep)

        # ── Step 3: Mating selection from archive ────────
        archive_fitness = spea2_fitness_assignment(archive)

        mating_pool = [
            spea2_tournament_select(archive, archive_fitness)
            for _ in range(pop_size)
        ]

        # ── Step 4: Crossover + Mutation → offspring ─────
        offspring = []
        while len(offspring) < pop_size:
            p1, p2 = random.sample(mating_pool, 2)
            if random.random() < CROSSOVER_RATE:
                c1, c2 = crossover_hybrid(p1, p2, batches, tt_dict, arc_lookup)
            else:
                c1 = random_initial_individual(batches, path_lib)
                c2 = random_initial_individual(batches, path_lib)

            # ── 修改后的正确代码 ──
            if random.random() < CROSSOVER_RATE:
                c1, c2 = crossover_hybrid(p1, p2, batches, tt_dict, arc_lookup)
            else:
                c1 = random_initial_individual(batches, path_lib)
                c2 = random_initial_individual(batches, path_lib)

            if random.random() < MUTATION_RATE:
                op, ok = mutate_fixed(c1, batches, path_lib, tt_dict, arc_lookup,
                                    arcs, waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
                          **eval_kwargs)
                mut_tracker["attempt"][op][gen] += 1

            if random.random() < MUTATION_RATE:
                op, ok = mutate_fixed(c2, batches, path_lib, tt_dict, arc_lookup,
                                      arcs, waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
                                      **eval_kwargs)
                mut_tracker["attempt"][op][gen] += 1

            repair_missing_allocations(c1, batches, path_lib)
            repair_missing_allocations(c2, batches, path_lib)
            evaluate_individual(c1, batches, arcs, tt_dict,     # ← 唯一一次评估
                                waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
                                **eval_kwargs)
            evaluate_individual(c2, batches, arcs, tt_dict,     # ← 唯一一次评估
                                waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
                                **eval_kwargs)
            offspring.extend([c1, c2])

        # ── New population = offspring (SPEA2 style) ─────
        population = offspring[:pop_size]

        # ── Feasibility boost ─────────────────────────────
        cur_feas = sum(1 for i in archive if i.feasible)
        boost_triggered = boost_new_feas = 0
        if cur_feas < MIN_FEASIBLE_SOLUTIONS:
            boost_triggered = 1
            archive, boost_new_feas = feasibility_boost(
                archive, batches, path_lib, tt_dict, arc_lookup,
                arcs,
                waiting_cost_per_teu_h, wait_emis_g_per_teu_h, eval_kwargs)
            after = sum(1 for i in archive if i.feasible)
            print(f"  ⚡ [BOOST] {cur_feas} < {MIN_FEASIBLE_SOLUTIONS} "
                  f"→ after boost: {after} (+{boost_new_feas})")

        boost_trigger_hist.append(boost_triggered)
        boost_new_feasible_hist.append(boost_new_feas)

    # ── Final archive → Pareto front ─────────────────────
    pareto = unique_individuals_by_objectives(
        [i for i in archive if i.feasible], tol=1e-3)

    total_t = time.perf_counter() - _run_start
    print(f"\n{'='*72}")
    print(f"  [SPEA2] Run complete: {generations} gens, {total_t:.1f}s, "
          f"Pareto={len(pareto)}")
    print(f"  ⚡ Boost: {sum(boost_trigger_hist)} gens triggered, "
          f"{sum(boost_new_feasible_hist)} new feasible")
    print(f"{'='*72}")

    return (population, pareto, batches,
            front_hist_objs,
            feasible_ratio_hist, feasible_ratio_strict_hist,
            vio_mean_hist, mut_tracker,
            boost_trigger_hist, boost_new_feasible_hist)


# ════════════════════════════════════════════════════════
# Scenario runner — calls SPEA2
# ════════════════════════════════════════════════════════

def run_scenario(
    scenario_id: str, K: int, alpha: float,
    node_names, node_region, node_hold_cost, node_proc_cost,
    arcs, timetables, raw_batches,
    waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
    carbon_tax_map, emission_factor_map, mode_speeds_map,
    trans_map, border_delay_map,
    pop_size:    int = 125,
    generations: int = 400,
    runs:        int = 30,
    base_seed:   int = 1000,
    output_dir:  str = ".",
) -> dict:
    out_dir = FSPath(output_dir) / scenario_id
    out_dir.mkdir(parents=True, exist_ok=True)

    tt_dict    = build_timetable_dict(timetables)
    arc_lookup = build_arc_lookup(arcs)

    run_front_hist, run_fr, run_frs      = [], [], []
    run_vio_mean, run_rows, run_paretos  = [], [], []
    run_batches_list, mut_runs           = [], []
    run_boost_trigger_hist               = []
    run_boost_new_feas_hist              = []

    for run_id in range(runs):
        seed = base_seed + run_id
        random.seed(seed); np.random.seed(seed)

        batches  = prepare_batches(raw_batches, K=K, alpha=alpha,
                                   node_region=node_region, random_seed=seed)
        path_lib = build_path_library(
            node_names, node_region, arcs, batches, tt_dict, arc_lookup)

        t0 = time.perf_counter()
        (pop, pareto, _, front_hist, fr_hist, frs_hist,
         vio_hist, mut_tracker, boost_trigger, boost_new_feas
        ) = run_spea2_analytics(
            _preloaded=(
                node_names, node_region, node_hold_cost, node_proc_cost,
                arcs, timetables, batches,
                waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
                carbon_tax_map, emission_factor_map, mode_speeds_map,
                trans_map, border_delay_map, path_lib,
            ),
            pop_size=pop_size, generations=generations,
        )
        runtime_s = float(time.perf_counter() - t0)

        run_front_hist.append(front_hist)
        run_fr.append(fr_hist);  run_frs.append(frs_hist)
        run_vio_mean.append(vio_hist)
        run_paretos.append(pareto); run_batches_list.append(batches)
        mut_runs.append(mut_tracker)
        run_boost_trigger_hist.append(boost_trigger)
        run_boost_new_feas_hist.append(boost_new_feas)

        run_rows.append({
            "scenario": scenario_id, "run_id": run_id, "seed": seed,
            "K": K, "alpha": alpha,
            "runtime_s": runtime_s,
            "final_feasible_ratio_soft":   float(fr_hist[-1])  if fr_hist  else 0.0,
            "final_feasible_ratio_strict": float(frs_hist[-1]) if frs_hist else 0.0,
            "final_pareto_size": int(len(pareto)),
            "boost_gens_triggered": int(sum(boost_trigger)),
        })
        print(f"  [{scenario_id}] Run {run_id:02d} | "
              f"Pareto={len(pareto)} | {runtime_s:.1f}s | "
              f"FeasSoft={fr_hist[-1]:.1%}")

    # ── Metrics ─────────────────────────────────────────
    P_star = build_P_star_fast(run_front_hist)
    if P_star:
        P_arr = np.array(P_star, dtype=float)
        mins, maxs = np.min(P_arr, axis=0), np.max(P_arr, axis=0)
    else:
        mins, maxs = np.zeros(3), np.ones(3)

    hv_calc = HypervolumeCalculator(
        ref_point=HV_REF_NORM, num_samples=HV_SAMPLES, seed=HV_MC_SEED)
    Pn = normalize_points(P_star, mins, maxs) if P_star else []

    hv_runs, igd_runs, sp_runs = [], [], []
    for r in range(runs):
        hv_h, igd_h, sp_h          = [], [], []
        last_hv, last_igd, last_sp = 0.0, float("inf"), 0.0
        for gi, gf in enumerate(run_front_hist[r]):
            if gi % HV_EVERY == 0:
                pts = [tuple(x) for x in _finite_points_array(gf)]
                An  = clip_points(normalize_points(pts, mins, maxs),
                                  HV_REF_NORM) if pts else []
                last_hv = hv_calc.calculate_points(An) if An else 0.0
            if gi % METRIC_EVERY == 0:
                pts     = [tuple(x) for x in _finite_points_array(gf)]
                An      = normalize_points(pts, mins, maxs) if pts else []
                last_igd = igd_plus(Pn, An) if (Pn and An) else float("inf")
                last_sp  = spacing_metric(An) if An else 0.0
            hv_h.append(last_hv); igd_h.append(last_igd); sp_h.append(last_sp)
        hv_runs.append(hv_h); igd_runs.append(igd_h); sp_runs.append(sp_h)

    hv_arr  = np.array(hv_runs,  dtype=float)
    igd_arr = np.array(igd_runs, dtype=float)
    sp_arr  = np.array(sp_runs,  dtype=float)
    fr_arr  = np.array(run_fr,   dtype=float)
    frs_arr = np.array(run_frs,  dtype=float)
    gen_arr = np.arange(generations)

    hv_mean, hv_std   = np.mean(hv_arr, 0),  np.std(hv_arr,  0)
    igd_mean, igd_std = np.mean(igd_arr, 0),  np.std(igd_arr, 0)
    sp_mean,  sp_std  = np.mean(sp_arr, 0),   np.std(sp_arr,  0)
    fr_mean,  fr_std  = np.mean(fr_arr, 0),   np.std(fr_arr,  0)
    frs_mean, frs_std = np.mean(frs_arr, 0),  np.std(frs_arr, 0)

    vio_keys = ["miss_alloc","miss_tt","cap_excess","late_teu_h","wait_teu_h"]
    vio_dict = {}
    for k in vio_keys:
        mat = np.array([run_vio_mean[r].get(k, [0.0]*generations)
                        for r in range(runs)], dtype=float)
        vio_dict[k] = list(np.mean(mat, axis=0))

    best_run_idx = int(np.argmax(hv_arr[:, -1]))
    best_fh      = run_front_hist[best_run_idx]
    best_boost   = run_boost_trigger_hist[best_run_idx]
    best_pareto  = run_paretos[best_run_idx]
    best_batches = run_batches_list[best_run_idx]

    def _bmin(hist, idx):
        vals = []
        for fh in hist:
            a = _finite_points_array(fh)
            vals.append(float(np.min(a[:, idx])) if a.shape[0] > 0 else np.nan)
        return _ffill_nan(np.array(vals))

    export_metrics_csv(
        gen_arr=gen_arr.tolist(),
        hv_mean=hv_mean.tolist(), hv_std=hv_std.tolist(),
        igd_mean=igd_mean.tolist(), igd_std=igd_std.tolist(),
        sp_mean=sp_mean.tolist(),   sp_std=sp_std.tolist(),
        fr_mean=fr_mean.tolist(),   fr_std=fr_std.tolist(),
        frs_mean=frs_mean.tolist(), frs_std=frs_std.tolist(),
        min_cost_best=_bmin(best_fh, 0).tolist(),
        min_emis_best=_bmin(best_fh, 1).tolist(),
        min_time_best=_bmin(best_fh, 2).tolist(),
        boost_hist_best=best_boost,
        vio_mean_dict_mean=vio_dict,
        out_csv=str(out_dir / "metrics_per_generation.csv"),
    )
    export_run_summary_excel(run_rows, hv_arr, igd_arr, sp_arr,
                             out_xlsx=str(out_dir / "run_summary.xlsx"))

    # ── Plots ────────────────────────────────────────────
    plot_hv_curve(gen_arr, hv_mean, hv_std,
                  save=str(out_dir / "plot_HV.png"))
    plot_igd_curve(gen_arr, igd_mean, igd_std,
                   save=str(out_dir / "plot_IGDplus.png"))
    plot_spacing_curve(gen_arr, sp_mean, sp_std,
                       save=str(out_dir / "plot_Spacing.png"))
    plot_convergence_combined(gen_arr, hv_mean, hv_std, igd_mean, igd_std,
                              save=str(out_dir / "plot_Convergence.png"))
    plot_feasible_ratio_curve(gen_arr, fr_mean, fr_std, frs_mean, frs_std,
                               best_boost,
                               save=str(out_dir / "plot_FeasibleRatio.png"))
    plot_min_objectives(gen_arr, _bmin(best_fh, 0), _bmin(best_fh, 1),
                        _bmin(best_fh, 2),
                        save=str(out_dir / "plot_MinObjectives.png"))

    mc_m, mc_s = extract_min_obj_per_gen_allruns(run_front_hist, 0)
    me_m, me_s = extract_min_obj_per_gen_allruns(run_front_hist, 1)
    mt_m, mt_s = extract_min_obj_per_gen_allruns(run_front_hist, 2)
    plot_min_cost_trend(gen_arr, mc_m, mc_s,
                        save=str(out_dir / "plot_MinCost_trend.png"))
    plot_min_emission_trend(gen_arr, me_m, me_s,
                            save=str(out_dir / "plot_MinEmission_trend.png"))
    plot_min_time_trend(gen_arr, mt_m, mt_s,
                        save=str(out_dir / "plot_MinTime_trend.png"))

    psz_mat = extract_pareto_size_per_gen(run_front_hist)
    plot_pareto_size_trend(gen_arr,
                           np.mean(psz_mat, axis=0), np.std(psz_mat, axis=0),
                           psz_mat[best_run_idx],
                           save=str(out_dir / "plot_ParetoSize_trend.png"))

    tfc_mat = fr_arr * pop_size
    plot_total_solution_count(gen_arr,
                              np.mean(tfc_mat, axis=0), np.std(tfc_mat, axis=0),
                              tfc_mat[best_run_idx],
                              save=str(out_dir / "plot_SolutionCount_trend.png"))

    best_pts = unique_objective_tuples(
        [ind.objectives for ind in best_pareto if ind.feasible], tol=1e-9)
    plot_pareto_3d(best_pts, save=str(out_dir / "plot_Pareto3D.png"),
                   title=f"SPEA2 {scenario_id} Pareto (K={K}, α={alpha})")
    plot_pareto_2d_projections(best_pts, save=str(out_dir / "plot_Pareto2D.png"))

    all_pts = [[ind.objectives for ind in run_paretos[r] if ind.feasible]
               for r in range(runs)]
    plot_pareto_2d_allruns(all_pts, best_run_idx=best_run_idx,
                           save=str(out_dir / "plot_Pareto2D_allruns.png"))
    plot_pareto_3d_allruns(all_pts, best_run_idx=best_run_idx,
                           save=str(out_dir / "plot_Pareto3D_allruns.png"))
    plot_objectives_scatter_by_run(all_pts,
                                   save=str(out_dir / "plot_ParetoByRun.png"))
    plot_summary_metrics_table(hv_arr, igd_arr, sp_arr, fr_arr, frs_arr,
                               run_rows, save=str(out_dir / "plot_SummaryTable.png"))

    prob_runs = [np.vstack([tr["prob"][op] for op in OPS]).astype(float)
                 for tr in mut_runs]
    prob_mean = np.mean(np.stack(prob_runs), axis=0)
    plot_mutation_adaptive_prob(gen_arr, prob_mean,
                                save=str(out_dir / "plot_MutationProb.png"))
    plot_runtime_per_run(run_rows,
                         save=str(out_dir / "plot_Runtime.png"))

    if best_pareto:
        save_pareto_solutions(best_pareto, best_batches,
                              str(out_dir / "result.txt"))
        export_pareto_points_json(best_pareto, best_batches,
                                  str(out_dir / "pareto_points.json"))

    # ── Export front_hist for cross-algorithm comparison ─
    import json as _json_fh
    _front_hist_export = []
    for _r_idx in range(len(run_front_hist)):
        _run_hist = []
        for _gen_front in run_front_hist[_r_idx]:
            _run_hist.append([list(obj) for obj in _gen_front])
        _front_hist_export.append(_run_hist)
    with open(str(out_dir / "spea2_front_hist.json"), "w", encoding="utf-8") as _f:
        _json_fh.dump(_front_hist_export, _f)
    print(f"[EXPORT] {out_dir}/spea2_front_hist.json saved.")

    # ── Return summary row ───────────────────────────────
    def _s(arr):
        f = arr[np.isfinite(arr)]
        return ((float(np.mean(f)), float(np.std(f)),
                 float(np.min(f)),  float(np.max(f)))
                if len(f) else (np.nan,)*4)

    hv_m,  hv_s,  hv_lo,  hv_hi  = _s(hv_arr[:,  -1])
    igd_m, igd_s, igd_lo, igd_hi  = _s(igd_arr[:, -1])
    sp_m,  sp_s,  sp_lo,  sp_hi   = _s(sp_arr[:,  -1])
    rt = np.array([r["runtime_s"] for r in run_rows])

    return {
        "scenario":          scenario_id,
        "K":                 K,
        "alpha":             alpha,
        "HV_mean":           hv_m,   "HV_std":  hv_s,
        "HV_min":            hv_lo,  "HV_max":  hv_hi,
        "IGD_mean":          igd_m,  "IGD_std": igd_s,
        "IGD_min":           igd_lo, "IGD_max": igd_hi,
        "Sp_mean":           sp_m,   "Sp_std":  sp_s,
        "Sp_min":            sp_lo,  "Sp_max":  sp_hi,
        "FeasSoft_mean":     float(np.mean(fr_arr[:,  -1])),
        "FeasStrict_mean":   float(np.mean(frs_arr[:, -1])),
        "Pareto_size_mean":  float(np.mean([r["final_pareto_size"] for r in run_rows])),
        "Pareto_size_std":   float(np.std ([r["final_pareto_size"] for r in run_rows])),
        "Runtime_mean_s":    float(np.mean(rt)),
        "Runtime_std_s":     float(np.std(rt)),
    }


# ════════════════════════════════════════════════════════
# Main — iterate over all scenarios with SPEA2
# ════════════════════════════════════════════════════════

OUTPUT_ROOT = "spea2_scenario_results"

if __name__ == "__main__":
    DATA_FILE         = "data.xlsx"
    POP_SIZE          = 125
    GENERATIONS       = 400
    RUNS_PER_SCENARIO = 30

    print("=" * 65)
    print("  SPEA2 Multi-Scenario Validation  (Table V, S0–S11)")
    print(f"  pop={POP_SIZE}  gens={GENERATIONS}  runs/scenario={RUNS_PER_SCENARIO}")
    print("=" * 65)

    print("\n[INIT] Loading network data...")
    (node_names, node_region,
     node_hold_cost, node_proc_cost,
     arcs, timetables, raw_batches,
     waiting_cost_per_teu_h, wait_emis_g_per_teu_h,
     carbon_tax_map, emission_factor_map, mode_speeds_map,
     trans_map, border_delay_map) = load_network_from_extended(DATA_FILE)
    pathlib.Path(OUTPUT_ROOT).mkdir(parents=True, exist_ok=True)
    all_summary_rows = []

    for sid, cfg in SCENARIO_TABLE.items():
        K, alpha = cfg["K"], cfg["alpha"]
        print(f"\n{'='*65}")
        print(f"  SCENARIO {sid}  |  K={K}  α={alpha}  (SPEA2)")
        print(f"{'='*65}")

        row = run_scenario(
            scenario_id=sid, K=K, alpha=alpha,
            node_names=node_names, node_region=node_region,
            node_hold_cost=node_hold_cost, node_proc_cost=node_proc_cost,
            arcs=arcs, timetables=timetables, raw_batches=raw_batches,
            waiting_cost_per_teu_h=waiting_cost_per_teu_h,
            wait_emis_g_per_teu_h=wait_emis_g_per_teu_h,
            carbon_tax_map=carbon_tax_map,
            emission_factor_map=emission_factor_map,
            mode_speeds_map=mode_speeds_map,
            trans_map=trans_map,
            border_delay_map=border_delay_map,
            pop_size=POP_SIZE, generations=GENERATIONS,
            runs=RUNS_PER_SCENARIO, output_dir=OUTPUT_ROOT,
        )
        all_summary_rows.append(row)
        print(f"  [{sid}] HV={row['HV_mean']:.4f}±{row['HV_std']:.4f}"
              f"  IGD+={row['IGD_mean']:.4f}"
              f"  FeasSoft={row['FeasSoft_mean']:.1%}"
              f"  Runtime={row['Runtime_mean_s']:.0f}s")

    # ── Cross-scenario summary Excel ──────────────────────
    df_sum   = pd.DataFrame(all_summary_rows)
    sum_path = FSPath(OUTPUT_ROOT) / "scenario_summary.xlsx"

    with pd.ExcelWriter(sum_path, engine="openpyxl") as writer:
        df_sum.to_excel(writer, sheet_name="ScenarioSummary", index=False)
        ws = writer.sheets["ScenarioSummary"]
        try:
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            hfill = PatternFill("solid", fgColor="1F4E79")
            hfont = Font(color="FFFFFF", bold=True)
            for ci, col in enumerate(df_sum.columns, 1):
                c = ws.cell(row=1, column=ci)
                c.fill = hfill; c.font = hfont
                c.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(ci)].width = max(len(col)+3, 12)
            ws.freeze_panes = "A2"
            for ri, rv in enumerate(df_sum.itertuples(), start=2):
                if rv.scenario == "S0":
                    for ci in range(1, len(df_sum.columns)+1):
                        ws.cell(row=ri, column=ci).fill = \
                            PatternFill("solid", fgColor="FFF2CC")
        except Exception:
            pass
    print(f"\n[EXPORT] Scenario summary → {sum_path}")

    # ── Cross-scenario comparison plots ───────────────────
    sids   = df_sum["scenario"].tolist()
    x      = range(len(sids))
    colors = ["#1976D2" if s == "S0" else "#E53935" for s in sids]

    fig, axes = plt.subplots(1, 3, figsize=(16, 5), dpi=180)
    for ax, (cm, cs, yl, ttl) in zip(axes, [
        ("HV_mean",  "HV_std",  "HV (normalised, ↑)",
         "SPEA2: HV across Scenarios"),
        ("IGD_mean", "IGD_std", "IGD+ (↓)",
         "SPEA2: IGD+ across Scenarios"),
        ("Sp_mean",  "Sp_std",  "Spacing (↓)",
         "SPEA2: Spacing across Scenarios"),
    ]):
        ax.bar(x, df_sum[cm], color=colors, alpha=0.82,
               yerr=df_sum[cs], capsize=4)
        ax.set_xticks(list(x))
        ax.set_xticklabels(sids, rotation=45, ha="right")
        ax.set_ylabel(yl); ax.set_title(ttl)
        ax.grid(axis="y", ls=":", alpha=0.5)
    fig.suptitle("SPEA2 Scenario Comparison  (S0=Baseline blue; Stress Tests red)",
                 fontsize=12, fontweight="bold")
    plt.tight_layout()
    cmp_path = FSPath(OUTPUT_ROOT) / "plot_ScenarioComparison.png"
    plt.savefig(cmp_path); plt.close()
    print(f"[PLOT] {cmp_path}")

    # ── Feasible-ratio heatmap ────────────────────────────
    pivot = df_sum.pivot_table(index="K", columns="alpha",
                               values="FeasSoft_mean")
    fig2, ax2 = plt.subplots(figsize=(6, 4), dpi=150)
    im = ax2.imshow(pivot.values, cmap="RdYlGn", vmin=0, vmax=1, aspect="auto")
    ax2.set_xticks(range(len(pivot.columns)))
    ax2.set_xticklabels([f"α={c}" for c in pivot.columns])
    ax2.set_yticks(range(len(pivot.index)))
    ax2.set_yticklabels([f"K={r}" for r in pivot.index])
    plt.colorbar(im, ax=ax2, label="Feasible Ratio (soft)")
    for i in range(len(pivot.index)):
        for j in range(len(pivot.columns)):
            v = pivot.values[i, j]
            ax2.text(j, i, f"{v:.2f}", ha="center", va="center",
                     color="black" if 0.3 < v < 0.85 else "white", fontsize=9)
    ax2.set_title("SPEA2: Feasible Ratio Heatmap  (K × α)")
    plt.tight_layout()
    hm_path = FSPath(OUTPUT_ROOT) / "plot_FeasHeatmap.png"
    plt.savefig(hm_path); plt.close()
    print(f"[PLOT] {hm_path}")

    # ── HV trend: K=20 group ─────────────────────────────
    k20_ids  = ["S0", "S4", "S5"]
    k20_rows = df_sum[df_sum["scenario"].isin(k20_ids)]
    fig3, ax3 = plt.subplots(figsize=(7, 4), dpi=160)
    for _, r in k20_rows.iterrows():
        ax3.bar(r["scenario"], r["HV_mean"],
                yerr=r["HV_std"], capsize=5,
                label=f"α={r['alpha']}", alpha=0.85)
    ax3.set_ylabel("HV (normalised)")
    ax3.set_title("SPEA2: HV K=20 group (varying α)")
    ax3.legend(); ax3.grid(axis="y", ls=":", alpha=0.5)
    plt.tight_layout()
    plt.savefig(FSPath(OUTPUT_ROOT) / "plot_HV_K20group.png")
    print(f"[PLOT] {FSPath(OUTPUT_ROOT) / 'plot_HV_K20group.png'}")

    print("\n" + "=" * 65)
    print("  SPEA2: ALL SCENARIOS COMPLETE")
    print(f"  Results in: {OUTPUT_ROOT}/")
    print("  Key outputs:")
    print("    scenario_summary.xlsx       — cross-scenario metrics table")
    print("    plot_ScenarioComparison.png — HV / IGD+ / Spacing bar chart")
    print("    plot_FeasHeatmap.png        — feasible ratio K×α heatmap")
    print("    <Sn>/metrics_per_generation.csv")
    print("    <Sn>/run_summary.xlsx")
    print("    <Sn>/result.txt  +  pareto_points.json")
    print("    <Sn>/plot_*.png  (12+ figures per scenario)")
    print("=" * 65)